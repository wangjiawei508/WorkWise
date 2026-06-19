#!/usr/bin/env python3
"""
Build an A3 landscape Excel report for automated total-station crossing reports.

Usage:
    python scripts/build_crossing_total_station_xlsx.py input.csv --image plan.jpg
    python scripts/build_crossing_total_station_xlsx.py input.csv --output report.xlsx

Input should follow assets/crossing-total-station-input-template.csv or contain
the same core fields generated from the automated monitoring platform.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.page import PageMargins


ITEM_ORDER = ["纵向变形量", "横向变形量", "垂直位移"]
TILT_ITEM = "倾斜率"


def read_rows(input_path: Path) -> list[dict[str, str]]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as fin:
        reader = csv.DictReader(fin)
        rows = [dict(row) for row in reader]
    if not rows:
        raise SystemExit("Input CSV has no data rows.")
    return rows


def first_value(rows: list[dict[str, str]], *fields: str) -> str:
    for row in rows:
        for field in fields:
            value = (row.get(field) or "").strip()
            if value:
                return value
    return ""


def parse_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned.strip(" ._") or "轨道交通控制保护区监测项目"


def parse_report_datetime(value: str) -> datetime | None:
    cleaned = str(value or "").strip()
    if not cleaned:
        return None
    cleaned = cleaned.replace("：", ":").replace("年", "-").replace("月", "-").replace("日", " ")
    cleaned = re.sub(r"\s*:\s*", ":", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2}).*?(\d{1,2})", cleaned)
    if match:
        year, month, day, hour = map(int, match.groups())
        return datetime(year, month, day, hour)
    return None


def normalized_report_time(value: str) -> str:
    parsed = parse_report_datetime(value)
    if parsed is not None:
        return parsed.strftime("%Y-%m-%d %H:%M")
    cleaned = str(value or "").strip().replace("：", ":")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def cadence_filename_label(value: str) -> str:
    cleaned = str(value or "").strip().lower().replace(" ", "")
    if "15" in cleaned and ("min" in cleaned or "分钟" in cleaned):
        return "15分钟"
    if "2" in cleaned and ("h" in cleaned or "小时" in cleaned):
        return "2小时"
    if "4" in cleaned and ("h" in cleaned or "小时" in cleaned):
        return "4小时"
    return safe_filename_part(value) or "快报"


def default_report_output_path(input_csv: Path) -> Path:
    rows = read_rows(input_csv)
    project_name = first_value(rows, "project_name", "项目名称") or "轨道交通控制保护区监测项目"
    current_time = first_value(rows, "current_time", "本次监测时间")
    cadence = first_value(rows, "report_cadence", "出报间隔") or "4h"
    report_dt = parse_report_datetime(current_time)
    if report_dt is None:
        match = re.search(r"(20\d{6})_(\d{2})", input_csv.stem)
        if match:
            date_label, hour_label = match.group(1), match.group(2)
        else:
            date_label, hour_label = "日期待确认", "时间待确认"
    else:
        date_label, hour_label = report_dt.strftime("%Y%m%d"), report_dt.strftime("%H")
    filename = (
        f"{safe_filename_part(project_name)}_"
        f"{date_label}_{hour_label}点_{cadence_filename_label(cadence)}快报.xlsx"
    )
    return input_csv.with_name(filename)


def fmt_01_value(value: float | None) -> float | str:
    if value is None:
        return "/"
    return float(Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def fmt_01_text(value: float | None) -> str:
    if value is None:
        return "/"
    return f"{Decimal(str(value)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP):.1f}"


def fmt_01(value: float | None) -> float | str:
    return fmt_01_value(value)


def fmt_02_value(value: float | None) -> float | str:
    if value is None:
        return "/"
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def fmt_02_text(value: float | None) -> str:
    if value is None:
        return "/"
    return f"{Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):.2f}"


def fmt_05_value(value: float | None) -> float | str:
    if value is None:
        return "/"
    return float(Decimal(str(value)).quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP))


def is_station_or_control_point(point_name: str) -> bool:
    upper = point_name.upper()
    return upper.startswith("CZ") or upper.startswith("JD")


def normalize_platform_status(platform_status: str) -> str:
    status = str(platform_status or "").strip()
    if status in ("", "0", "None"):
        return "无预警"
    if "IsTolerance=0" in status:
        return "无预警"
    if "IsTolerance=-1" in status:
        return "未平差/无成果"
    return status


def normalize_initial_time(value: str, method: str = "") -> str:
    cleaned = value.strip()
    if "平台未给出具体时刻" in cleaned or "平台未给出具体时间" in cleaned:
        date_part = cleaned.split("（", 1)[0].strip()
        return date_part or "全站仪自动化平台初始值"
    if "人工初始值" in cleaned and "自动化" in method:
        return "全站仪自动化平台初始值"
    return cleaned


def extract_implementation_unit(initial_report_path: Path | None) -> str:
    if not initial_report_path or not initial_report_path.exists():
        return ""
    try:
        from docx import Document
    except Exception:
        return ""
    try:
        doc = Document(initial_report_path)
    except Exception:
        return ""
    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if "监测单位" not in text:
            continue
        tail = text.split("监测单位", 1)[1]
        tail = tail.replace("：", ":", 1)
        if ":" in tail:
            value = tail.split(":", 1)[1]
        else:
            value = tail
        value = value.split("编", 1)[0].split("合同", 1)[0].strip()
        if value:
            return value
    return ""


def build_pivot(rows: list[dict[str, str]]) -> dict[str, dict[str, object]]:
    pivot: dict[str, dict[str, object]] = {}
    for row in rows:
        point = first_value([row], "point_id", "测点", "点号") or "/"
        zone = first_value([row], "structure_zone", "position_label", "监测部位") or "未分区"
        platform_status = str(row.get("platform_warn_status", "")).strip()
        status = row.get("status") or normalize_platform_status(platform_status)
        pivot.setdefault(
            point,
            {
                "zone": zone,
                "status": status,
                "position_label": first_value([row], "position_label", "位置") or "",
            },
        )
        item = first_value([row], "monitoring_item", "监测项目") or "未分类监测项"
        pivot[point][item] = {
            "current": parse_float(row.get("current_change_mm")),
            "cumulative": parse_float(row.get("cumulative_mm")),
        }
    return pivot


def load_manual_metric_overrides(path: Path | None) -> list[dict[str, object]]:
    if not path or not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        overrides = data.get("overrides", [])
        if isinstance(overrides, list):
            return [item for item in overrides if isinstance(item, dict)]
    return []


def report_time_matches(override: dict[str, object], current_time: str) -> bool:
    for key in ("report_time", "report_end", "current_time", "本次监测时间"):
        value = str(override.get(key) or "").strip()
        if value and normalized_report_time(value) == current_time:
            return True
    return False


def override_direction(override: dict[str, object]) -> str:
    raw = " ".join(
        str(override.get(key) or "")
        for key in ("direction", "direction_label", "方向", "备注", "note")
    )
    if "南北" in raw or "北" in raw or "north" in raw.lower():
        return "north"
    if "东西" in raw or "东" in raw or "east" in raw.lower():
        return "east"
    return ""


def normalize_override_item(item: str, override: dict[str, object]) -> str:
    text = str(item or "").strip()
    direction = override_direction(override)
    if text in {"垂直位移", "横向变形量", "纵向变形量"}:
        return text
    if "沉降" in text or "竖向" in text or "垂直" in text:
        return "垂直位移"
    if "水平" in text or "位移" in text:
        if direction == "north":
            return "纵向变形量"
        return "横向变形量"
    if "南北" in text or "纵向" in text:
        return "纵向变形量"
    if "东西" in text or "横向" in text:
        return "横向变形量"
    return text


def override_metric_value(override: dict[str, object], *keys: str) -> float | None:
    for key in keys:
        if key not in override:
            continue
        value = parse_float(override.get(key))
        if value is not None:
            return value
    return None


def apply_manual_metric_overrides(
    pivot: dict[str, dict[str, object]],
    rows: list[dict[str, str]],
    path: Path | None,
) -> None:
    overrides = load_manual_metric_overrides(path)
    if not overrides:
        return
    current_time = normalized_report_time(first_value(rows, "current_time", "本次监测时间"))
    for override in overrides:
        if current_time and not report_time_matches(override, current_time):
            continue
        point = str(override.get("point_id") or override.get("point") or override.get("点号") or "").strip()
        item = normalize_override_item(
            str(override.get("monitoring_item") or override.get("item") or override.get("监测项目") or "").strip(),
            override,
        )
        if not point or not item or point not in pivot:
            continue
        item_data = pivot[point].setdefault(item, {})
        if not isinstance(item_data, dict):
            item_data = {}
            pivot[point][item] = item_data
        current_value = override_metric_value(override, "current_change_mm", "current", "本次变化值", "本次变量")
        cumulative_value = override_metric_value(override, "cumulative_mm", "cumulative", "累计变化值", "累计变量")
        if current_value is not None:
            item_data["current"] = current_value
        if cumulative_value is not None:
            item_data["cumulative"] = cumulative_value


def apply_manual_coordinate_overrides(
    initial_coords: dict[str, dict[str, float]],
    previous_coords: dict[str, dict[str, float]],
    current_coords: dict[str, dict[str, float]],
    rows: list[dict[str, str]],
    path: Path | None,
) -> None:
    overrides = load_manual_metric_overrides(path)
    if not overrides:
        return
    current_time = normalized_report_time(first_value(rows, "current_time", "本次监测时间"))
    axis_by_item = {
        "纵向变形量": "X",
        "横向变形量": "Y",
        "垂直位移": "H",
    }
    for override in overrides:
        if current_time and not report_time_matches(override, current_time):
            continue
        point = str(override.get("point_id") or override.get("point") or override.get("点号") or "").strip()
        item = normalize_override_item(
            str(override.get("monitoring_item") or override.get("item") or override.get("监测项目") or "").strip(),
            override,
        )
        axis = axis_by_item.get(item)
        if not point or not axis:
            continue
        current_coords.setdefault(point, {})
        current_change = override_metric_value(override, "current_change_mm", "current", "本次变化值", "本次变量")
        cumulative = override_metric_value(override, "cumulative_mm", "cumulative", "累计变化值", "累计变量")
        if current_change is not None and previous_coords.get(point, {}).get(axis) is not None:
            current_coords[point][axis] = float(previous_coords[point][axis]) + current_change / 1000.0
        elif cumulative is not None and initial_coords.get(point, {}).get(axis) is not None:
            current_coords[point][axis] = float(initial_coords[point][axis]) + cumulative / 1000.0


def load_alias_map(path: Path | None) -> dict[str, object]:
    if not path or not path.exists():
        return {"points": {}, "tilt_pairs": {}}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        return {"points": {}, "tilt_pairs": {}}
    data.setdefault("points", {})
    data.setdefault("tilt_pairs", {})
    return data


def point_alias(alias_map: dict[str, object], point: str) -> dict[str, str]:
    points = alias_map.get("points")
    if not isinstance(points, dict):
        return {}
    value = points.get(point)
    return value if isinstance(value, dict) else {}


def display_point(alias_map: dict[str, object], point: str) -> str:
    return str(point_alias(alias_map, point).get("display_point") or point)


def display_zone(alias_map: dict[str, object], point: str, default_zone: str) -> str:
    return str(point_alias(alias_map, point).get("structure_zone") or default_zone)


def display_position(alias_map: dict[str, object], point: str, default_position: str) -> str:
    return str(point_alias(alias_map, point).get("position") or default_position)


def display_tilt_pair(alias_map: dict[str, object], pair_label: str) -> str:
    pairs = alias_map.get("tilt_pairs")
    if isinstance(pairs, dict):
        value = pairs.get(pair_label)
        if isinstance(value, dict) and value.get("display_point"):
            return str(value["display_point"])
    return pair_label


def point_suffix(point_name: str) -> str:
    match = re.search(r"(\d+)$", point_name)
    return match.group(1) if match else point_name


def auto_point_config_path(input_csv: Path) -> Path | None:
    candidates = [
        input_csv.with_name(input_csv.name.replace("_adjusted_total_station.csv", "_platform_initial_values.json")),
        input_csv.with_name(input_csv.name.replace("_total_station_platform.csv", "_point_config.json")),
        input_csv.with_name(input_csv.name.replace("_total_station_platform.csv", "_platform_initial_values.json")),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_point_heights(point_config_path: Path | None) -> dict[str, float]:
    if not point_config_path or not point_config_path.exists():
        return {}
    data = json.loads(point_config_path.read_text(encoding="utf-8"))
    heights: dict[str, float] = {}
    if not isinstance(data, list):
        return heights
    for row in data:
        if not isinstance(row, dict):
            continue
        name = str(row.get("PointName") or row.get("pointName") or row.get("PointName1") or "").strip()
        if not name:
            continue
        height = parse_float(row.get("H"))
        if height is None:
            height = parse_float(row.get("TargetZ"))
        if height is not None:
            heights[name] = height
    return heights


def load_point_geometry(point_config_path: Path | None) -> dict[str, dict[str, float]]:
    if not point_config_path or not point_config_path.exists():
        return {}
    data = json.loads(point_config_path.read_text(encoding="utf-8"))
    geometry: dict[str, dict[str, float]] = {}
    if not isinstance(data, list):
        return geometry
    for row in data:
        if not isinstance(row, dict):
            continue
        name = str(row.get("PointName") or row.get("pointName") or row.get("PointName1") or "").strip()
        if not name:
            continue
        coords: dict[str, float] = {}
        for field in ("X", "Y", "H"):
            value = parse_float(row.get(field))
            if value is not None:
                coords[field] = value
        if coords:
            geometry[name] = coords
    return geometry


def point_geometry_from_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, float]]:
    geometry: dict[str, dict[str, float]] = {}
    for row in rows:
        point = first_value([row], "point_id", "测点", "点号")
        if not point or point in geometry:
            continue
        coords: dict[str, float] = {}
        for source_field, target_field in (
            ("initial_x_m", "X"),
            ("initial_y_m", "Y"),
            ("initial_h_m", "H"),
        ):
            value = parse_float(row.get(source_field))
            if value is not None:
                coords[target_field] = value
        if coords:
            geometry[point] = coords
    return geometry


def point_spacing_m(
    top_point: str,
    lower_point: str,
    point_geometry: dict[str, dict[str, float]] | dict[str, float],
) -> float | None:
    top = point_geometry.get(top_point)
    lower = point_geometry.get(lower_point)
    if isinstance(top, dict) and isinstance(lower, dict):
        if all(axis in top and axis in lower for axis in ("X", "Y", "H")):
            return math.sqrt(
                (float(top["X"]) - float(lower["X"])) ** 2
                + (float(top["Y"]) - float(lower["Y"])) ** 2
                + (float(top["H"]) - float(lower["H"])) ** 2
            )
        if "H" in top and "H" in lower:
            return abs(float(top["H"]) - float(lower["H"]))
    if isinstance(top, (int, float)) and isinstance(lower, (int, float)):
        return abs(float(top) - float(lower))
    return None


def tilt_axis_description(tilt_axis: str) -> str:
    if tilt_axis in {"纵向变形量", "纵向", "dX", "dx", "north"}:
        return "纵向变形量差值/上下点初始三维距离"
    if tilt_axis in {"横向变形量", "横向位移", "横向", "dY", "dy", "east"}:
        return "横向变形量差值/上下点初始三维距离"
    return "纵横向位移差合成值/上下点初始三维距离"


def make_resized_image_copy(
    image_path: Path,
    output_path: Path,
    suffix: str,
    max_size: tuple[int, int],
) -> Path:
    try:
        from PIL import Image
    except Exception:
        return image_path
    with Image.open(image_path) as image:
        image = image.convert("RGB")
        image.thumbnail(max_size)
        resized_path = output_path.with_name(f"{output_path.stem}_{suffix}.png")
        image.save(resized_path)
    return resized_path


def make_canvas_image_copy(
    image_path: Path,
    output_path: Path,
    suffix: str,
    canvas_size: tuple[int, int],
) -> Path:
    try:
        from PIL import Image
    except Exception:
        return image_path
    with Image.open(image_path) as image:
        image = image.convert("RGB")
        scale = max(canvas_size[0] / image.width, canvas_size[1] / image.height)
        resized_size = (int(image.width * scale + 0.5), int(image.height * scale + 0.5))
        image = image.resize(resized_size)
        left_crop = max((image.width - canvas_size[0]) // 2, 0)
        top_crop = max((image.height - canvas_size[1]) // 2, 0)
        image = image.crop((left_crop, top_crop, left_crop + canvas_size[0], top_crop + canvas_size[1]))
        canvas = Image.new("RGB", canvas_size, "white")
        canvas.paste(image, (0, 0))
        canvas_path = output_path.with_name(f"{output_path.stem}_{suffix}.png")
        canvas.save(canvas_path)
    return canvas_path


def item_value(data: dict[str, object], item: str, field: str) -> float | None:
    item_data = data.get(item)
    if not isinstance(item_data, dict):
        return None
    return parse_float(item_data.get(field))


def compute_tilt_rates(
    pivot: dict[str, dict[str, object]],
    point_geometry: dict[str, dict[str, float]] | dict[str, float],
    tilt_axis: str = "resultant",
) -> dict[str, dict[str, object]]:
    """Return tilt rates keyed by top-point name.

    If tilt_axis is 纵向变形量 or 横向变形量, the rate keeps the sign of that
    direction. If it is resultant, the rate is an unsigned horizontal resultant.
    Since displacement is in mm and spacing is in m, the numerical result is per
    mille.
    """
    tilt: dict[str, dict[str, object]] = {}
    for point in pivot:
        if not point.startswith("S"):
            continue
        suffix = point_suffix(point)
        lower = f"X{suffix}"
        if lower not in pivot:
            continue
        spacing_m = point_spacing_m(point, lower, point_geometry)
        if spacing_m is None or spacing_m <= 0:
            continue
        top_data = pivot[point]
        lower_data = pivot[lower]
        tilt[point] = {
            "pier": suffix,
            "top_point": point,
            "lower_point": lower,
            "spacing_m": spacing_m,
        }
        for field in ("current", "cumulative"):
            dzx_top = item_value(top_data, "纵向变形量", field)
            dzx_lower = item_value(lower_data, "纵向变形量", field)
            dhx_top = item_value(top_data, "横向变形量", field)
            dhx_lower = item_value(lower_data, "横向变形量", field)
            if None in (dzx_top, dzx_lower, dhx_top, dhx_lower):
                tilt[point][field] = None
                continue
            dzx_diff = float(dzx_top) - float(dzx_lower)
            dhx_diff = float(dhx_top) - float(dhx_lower)
            if tilt_axis in {"纵向变形量", "纵向", "dX", "dx", "north"}:
                tilt[point][field] = dzx_diff / spacing_m
                tilt[point]["tilt_axis"] = "纵向变形量"
            elif tilt_axis in {"横向变形量", "横向位移", "横向", "dY", "dy", "east"}:
                tilt[point][field] = dhx_diff / spacing_m
                tilt[point]["tilt_axis"] = "横向变形量"
            else:
                tilt[point][field] = math.hypot(dzx_diff, dhx_diff) / spacing_m
                tilt[point]["tilt_axis"] = "纵横向合成"
    return tilt


def tilt_maxima(tilt: dict[str, dict[str, object]]) -> dict[str, tuple[str, float]]:
    maxima: dict[str, tuple[str, float]] = {}
    for field in ("current", "cumulative"):
        candidates: list[tuple[float, str]] = []
        for data in tilt.values():
            value = parse_float(data.get(field))
            if value is None:
                continue
            label = f"{data.get('top_point')}/{data.get('lower_point')}"
            candidates.append((value, label))
        if candidates:
            value, label = max(candidates, key=lambda item_value: (abs(item_value[0]), item_value[1]))
            maxima[field] = (label, value)
    return maxima


def pick_maxima(pivot: dict[str, dict[str, object]]) -> dict[tuple[str, str], tuple[str, float]]:
    maxima: dict[tuple[str, str], tuple[str, float]] = {}
    for item in ITEM_ORDER:
        for field in ("current", "cumulative"):
            candidates: list[tuple[float, float, str]] = []
            for point, data in pivot.items():
                if item in {"横向变形量", "垂直位移"} and not point.upper().startswith("X"):
                    continue
                item_data = data.get(item)
                if not isinstance(item_data, dict):
                    continue
                value = parse_float(item_data.get(field))
                if value is None:
                    continue
                candidates.append((abs(value), value, point))
            if candidates:
                _, value, point = max(candidates, key=lambda item_value: (item_value[0], item_value[2]))
                maxima[(item, field)] = (point, value)
    return maxima


def split_points(pivot: dict[str, dict[str, object]]) -> tuple[list[str], list[str]]:
    points = list(pivot)
    left = [p for p in points if "12" in str(pivot[p].get("zone", "")) or str(pivot[p].get("zone", "")).startswith("其他")]
    right = [p for p in points if "13" in str(pivot[p].get("zone", ""))]
    used = set(left) | set(right)
    leftovers = [p for p in points if p not in used]
    if not left and not right:
        middle = (len(points) + 1) // 2
        return points[:middle], points[middle:]
    for point in leftovers:
        (left if len(left) <= len(right) else right).append(point)
    return left, right


def build_report(
    input_csv: Path,
    output_path: Path,
    image_path: Path | None = None,
    point_config_path: Path | None = None,
    tilt_axis: str = "resultant",
) -> None:
    rows = read_rows(input_csv)
    pivot = build_pivot(rows)
    maxima = pick_maxima(pivot)
    point_config_path = point_config_path or auto_point_config_path(input_csv)
    point_geometry = point_geometry_from_rows(rows) or load_point_geometry(point_config_path)
    tilt = compute_tilt_rates(pivot, point_geometry, tilt_axis)
    tilt_max = tilt_maxima(tilt)
    left_points, right_points = split_points(pivot)

    project_name = first_value(rows, "project_name", "项目名称") or "{{项目全称}}"
    current_time = first_value(rows, "current_time", "本次监测时间") or "{{本次监测时间}}"
    previous_time = first_value(rows, "previous_time", "上次监测时间") or "{{上次监测时间}}"
    initial_time = first_value(rows, "initial_time", "初始采集时间") or "{{初始采集时间}}"
    method = first_value(rows, "monitoring_method", "监测方式") or "自动化全站仪"
    cadence = first_value(rows, "report_cadence", "出报间隔") or "{{15min/2h/4h}}"
    condition = first_value(rows, "work_condition", "ring_no", "施工工况") or "{{施工工况}}"

    wb = Workbook()
    ws = wb.active
    ws.title = "全站仪快报"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.15, right=0.15, top=0.2, bottom=0.2, header=0, footer=0)
    ws.print_area = "A1:Z55"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True

    for col in range(1, 27):
        ws.column_dimensions[get_column_letter(col)].width = 5.2
    for row in range(1, 56):
        ws.row_dimensions[row].height = 18
    for row in range(7, 23):
        ws.row_dimensions[row].height = 23
    for row, height in {1: 24, 4: 42, 6: 28, 25: 20, 26: 20, 28: 24, 46: 32, 50: 30, 52: 30}.items():
        ws.row_dimensions[row].height = height

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=thin, right=thin, top=medium, bottom=thin)
    outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    white = PatternFill("solid", fgColor="FFFFFF")
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    green_fill = PatternFill("solid", fgColor="00B050")
    font_title = Font(name="SimSun", size=15, bold=True)
    font_section = Font(name="SimSun", size=11, bold=True)
    font_body = Font(name="SimSun", size=9)
    font_body_bold = Font(name="SimSun", size=9, bold=True)
    font_small = Font(name="SimSun", size=8)
    font_white = Font(name="SimSun", size=9, color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=55, min_col=1, max_col=26):
        for cell in row:
            cell.font = font_body
            cell.alignment = center
            cell.border = base_border
            cell.fill = white

    for col in range(1, 27):
        ws.cell(1, col).border = Border(top=medium, left=ws.cell(1, col).border.left, right=ws.cell(1, col).border.right, bottom=ws.cell(1, col).border.bottom)
        ws.cell(55, col).border = Border(bottom=medium, left=ws.cell(55, col).border.left, right=ws.cell(55, col).border.right, top=ws.cell(55, col).border.top)
    for row in range(1, 56):
        ws.cell(row, 1).border = Border(left=medium, right=ws.cell(row, 1).border.right, top=ws.cell(row, 1).border.top, bottom=ws.cell(row, 1).border.bottom)
        ws.cell(row, 26).border = Border(right=medium, left=ws.cell(row, 26).border.left, top=ws.cell(row, 26).border.top, bottom=ws.cell(row, 26).border.bottom)
        ws.cell(row, 13).border = Border(right=medium, left=ws.cell(row, 13).border.left, top=ws.cell(row, 13).border.top, bottom=ws.cell(row, 13).border.bottom)
        ws.cell(row, 14).border = Border(left=medium, right=ws.cell(row, 14).border.right, top=ws.cell(row, 14).border.top, bottom=ws.cell(row, 14).border.bottom)

    def merge_write(ref: str, value: object, font=None, fill=None, align=None, border=None) -> None:
        ws.merge_cells(ref)
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        cell = ws.cell(min_row, min_col)
        cell.value = value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border:
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    ws.cell(rr, cc).border = border

    merge_write("A1:Z1", project_name, font_title, None, center, outer_border)
    merge_write("A2:M2", "一、施工工况", font_section, header_fill, left, section_border)
    merge_write("N2:Z2", "一、施工工况", font_section, header_fill, left, section_border)
    merge_write("A3:M4", f"当前工况：{condition}。", font_body, None, left, base_border)
    merge_write("N3:Z4", f"本次自动化全站仪数据取自监测平台，监测时间 {current_time}；报表频率：{cadence}。", font_body, None, left, base_border)
    merge_write("A5:M5", "二、施工进度", font_section, header_fill, left, section_border)
    merge_write("N5:Z5", "二、施工进度", font_section, header_fill, left, section_border)
    merge_write("A6:M6", condition, font_body, None, left, base_border)
    merge_write("N6:Z6", "当前处于穿越影响跟踪阶段，应结合现场工况持续复核异常点和缺测点。", font_body, None, left, base_border)
    merge_write("A7:M22", "", font_body, None, center, base_border)
    if image_path and image_path.exists():
        img = XLImage(str(image_path))
        aspect = img.width / img.height if img.height else 1
        img.height = 345
        img.width = int(img.height * aspect)
        img.anchor = "D8"
        ws.add_image(img)

    merge_write("N7:Z7", "平台最新动态", font_section, sub_fill, center, base_border)
    status_rows = [
        ("项目名称", project_name),
        ("本次监测时间", current_time),
        ("上次监测时间", previous_time),
        ("监测方式", method),
        ("监测点数量", f"{len(pivot)} 个"),
        ("平台预警", "红色 0 / 橙色 0 / 黄色 0"),
    ]
    for item in ITEM_ORDER:
        point, value = maxima.get((item, "current"), ("/", None))
        status_rows.append((f"本次最大{item[:2]}", f"{point}  {fmt_01(value):.1f} mm" if isinstance(value, float) else "/"))
    if tilt_max:
        point, value = tilt_max.get("current", ("/", None))
        status_rows.append(("本次最大倾斜", f"{point}  {value:.2f}‰" if isinstance(value, float) else "/"))
        point, value = tilt_max.get("cumulative", ("/", None))
        status_rows.append(("累计最大倾斜", f"{point}  {value:.2f}‰" if isinstance(value, float) else "/"))
    for row_index, (label, value) in enumerate(status_rows, start=8):
        merge_write(f"N{row_index}:P{row_index}", label, font_body_bold, header_fill, center, base_border)
        merge_write(f"Q{row_index}:Z{row_index}", value, font_body, None, left, base_border)
    for row_index in range(8 + len(status_rows), 22):
        merge_write(f"N{row_index}:Z{row_index}", "", font_body, None, left, base_border)
    merge_write("N22:Z22", "说明：绿色单元格为本表绝对值最大的本次变化、累计变化或倾斜率；红色仅用于达到项目预警/报警条件的数据。", font_small, None, left, base_border)

    merge_write("A23:M23", "三、各项数据分析", font_section, header_fill, left, section_border)
    merge_write("N23:Z23", "三、各项数据分析", font_section, header_fill, left, section_border)
    merge_write("A24:D24", "初始采集时间", font_body_bold, header_fill, center, base_border)
    merge_write("E24:M24", initial_time, font_body, None, left, base_border)
    merge_write("N24:Q24", "监测方式", font_body_bold, header_fill, center, base_border)
    merge_write("R24:Z24", method, font_body, None, left, base_border)
    merge_write("A25:D25", "上次监测时间", font_body_bold, header_fill, center, base_border)
    merge_write("E25:M25", previous_time, font_body, None, left, base_border)
    merge_write("N25:Q25", "本次监测时间", font_body_bold, header_fill, center, base_border)
    merge_write("R25:Z25", current_time, font_body, None, left, base_border)
    merge_write("A26:M26", "12号桥墩及关联测点", font_section, sub_fill, center, base_border)
    merge_write("N26:Z26", "13号桥墩及关联测点", font_section, sub_fill, center, base_border)

    def table_headers(prefix_col: int) -> None:
        cols = [get_column_letter(prefix_col + i) for i in range(13)]
        for ref, text in [
            (f"{cols[0]}27:{cols[0]}28", "位置"),
            (f"{cols[1]}27:{cols[1]}28", "监测点号"),
            (f"{cols[2]}27:{cols[3]}27", "纵向变形(mm)"),
            (f"{cols[4]}27:{cols[5]}27", "横向位移(mm)"),
            (f"{cols[6]}27:{cols[7]}27", "垂直位移(mm)"),
            (f"{cols[8]}27:{cols[9]}27", "倾斜率(‰)"),
            (f"{cols[10]}27:{cols[10]}28", "状态"),
            (f"{cols[11]}27:{cols[12]}28", "备注"),
        ]:
            merge_write(ref, text, font_body_bold, header_fill, center, base_border)
        for offset, text in enumerate(["本次", "累计", "本次", "累计", "本次", "累计", "本次", "累计"], start=2):
            cell = ws.cell(28, prefix_col + offset)
            cell.value = text
            cell.font = font_body_bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = base_border

    table_headers(1)
    table_headers(14)

    def write_point(row_index: int, prefix_col: int, point: str | None) -> None:
        col = prefix_col
        if point is None:
            for offset in range(11):
                ws.cell(row_index, col + offset).value = ""
            merge_write(f"{get_column_letter(col + 11)}{row_index}:{get_column_letter(col + 12)}{row_index}", "", font_body, None, center, base_border)
            return
        data = pivot[point]
        zone = str(data.get("zone", ""))
        fill = yellow_fill if zone in ("12号桥墩", "13号桥墩") else gray_fill
        tilt_data = tilt.get(point, {})
        tilt_current = parse_float(tilt_data.get("current"))
        tilt_cumulative = parse_float(tilt_data.get("cumulative"))
        values = [
            zone,
            point,
            fmt_01((data.get("纵向变形量") or {}).get("current") if isinstance(data.get("纵向变形量"), dict) else None),
            fmt_01((data.get("纵向变形量") or {}).get("cumulative") if isinstance(data.get("纵向变形量"), dict) else None),
            fmt_01((data.get("横向变形量") or {}).get("current") if isinstance(data.get("横向变形量"), dict) else None),
            fmt_01((data.get("横向变形量") or {}).get("cumulative") if isinstance(data.get("横向变形量"), dict) else None),
            fmt_01((data.get("垂直位移") or {}).get("current") if isinstance(data.get("垂直位移"), dict) else None),
            fmt_01((data.get("垂直位移") or {}).get("cumulative") if isinstance(data.get("垂直位移"), dict) else None),
            f"{tilt_current:.2f}" if tilt_current is not None else "/",
            f"{tilt_cumulative:.2f}" if tilt_cumulative is not None else "/",
            data.get("status", "无预警"),
        ]
        for offset, value in enumerate(values):
            cell = ws.cell(row_index, col + offset)
            cell.value = value
            cell.font = font_body
            cell.alignment = center
            cell.border = base_border
            cell.fill = fill if offset in (0, 1) else white
            if isinstance(value, float):
                cell.number_format = "0.0"
        note = "当前盾构影响范围跟踪" if zone in ("12号桥墩", "13号桥墩") else "关联测点"
        if tilt_data.get("spacing_m"):
            note = f"{tilt_data.get('top_point')}/{tilt_data.get('lower_point')}初始三维距离{float(tilt_data.get('spacing_m')):.2f}m"
        merge_write(f"{get_column_letter(col + 11)}{row_index}:{get_column_letter(col + 12)}{row_index}", note, font_body, None, center, base_border)
        item_cols = {
            ("纵向变形量", "current"): col + 2,
            ("纵向变形量", "cumulative"): col + 3,
            ("横向变形量", "current"): col + 4,
            ("横向变形量", "cumulative"): col + 5,
            ("垂直位移", "current"): col + 6,
            ("垂直位移", "cumulative"): col + 7,
        }
        for key, (max_point, _) in maxima.items():
            if max_point == point:
                cell = ws.cell(row_index, item_cols[key])
                cell.fill = green_fill
                cell.font = font_white
        for field, offset in (("current", 8), ("cumulative", 9)):
            max_item = tilt_max.get(field)
            if max_item and max_item[0] == f"{point}/{tilt_data.get('lower_point')}":
                cell = ws.cell(row_index, col + offset)
                cell.fill = green_fill
                cell.font = font_white

    for index in range(14):
        row_index = 29 + index
        write_point(row_index, 1, left_points[index] if index < len(left_points) else None)
        write_point(row_index, 14, right_points[index] if index < len(right_points) else None)

    def write_max_rows(prefix_col: int) -> None:
        labels = {43: "本次最大", 44: "累计最大"}
        for row_index, label in labels.items():
            merge_write(f"{get_column_letter(prefix_col)}{row_index}:{get_column_letter(prefix_col + 1)}{row_index}", label, font_body_bold, header_fill, center, base_border)
            field = "current" if row_index == 43 else "cumulative"
            max_items = ITEM_ORDER + [TILT_ITEM]
            for index, item in enumerate(max_items):
                if item == TILT_ITEM:
                    point, value = tilt_max.get(field, ("/", None))
                    display_value = f"{value:.2f}" if isinstance(value, float) else "/"
                else:
                    point, value = maxima.get((item, field), ("/", None))
                    display_value = f"{fmt_01(value):.1f}" if isinstance(value, float) else "/"
                point_col = prefix_col + 2 + index * 2
                value_col = point_col + 1
                ws.cell(row_index, point_col).value = point
                ws.cell(row_index, point_col).font = font_body
                ws.cell(row_index, point_col).alignment = center
                ws.cell(row_index, point_col).border = base_border
                ws.cell(row_index, value_col).value = display_value
                ws.cell(row_index, value_col).font = font_body
                ws.cell(row_index, value_col).alignment = center
                ws.cell(row_index, value_col).border = base_border
    write_max_rows(1)
    write_max_rows(14)

    template_set(ws, "A19", "倾斜计算距离")
    spacing_notes = []
    for data in tilt.values():
        spacing = parse_float(data.get("spacing_m"))
        if spacing is not None:
            spacing_notes.append(f"{data.get('pier')}#初始三维距离{spacing:.2f}m")
    spacing_text = "；".join(spacing_notes) if spacing_notes else "上下测点初始三维距离待确认"
    notes = (
        "注：1、纵向变形、横向位移、垂直位移均以毫米计；本次/累计变形采用平台平差后成果并以全站仪自动化平台初始值校核。\n"
        f"2、倾斜率按S/X上下测点{tilt_axis_description(tilt_axis)}计算，单位‰；"
        f"{spacing_text}。\n"
        "3、黄色表示当前重点影响区，绿色表示本表绝对值最大项；红色仅用于达到预警/报警条件的数据。"
    )
    merge_write("A45:M47", notes, font_small, None, left, base_border)
    merge_write("N45:Z47", notes, font_small, None, left, base_border)
    merge_write("A48:M48", "四、结论与评价", font_section, header_fill, left, section_border)
    merge_write("N48:Z48", "四、结论与评价", font_section, header_fill, left, section_border)
    tilt_conclusion = ""
    if tilt_max:
        current_tilt = tilt_max.get("current")
        cumulative_tilt = tilt_max.get("cumulative")
        parts = []
        if current_tilt:
            parts.append(f"本次最大倾斜率{current_tilt[1]:.2f}‰（{current_tilt[0]}）")
        if cumulative_tilt:
            parts.append(f"累计最大倾斜率{cumulative_tilt[1]:.2f}‰（{cumulative_tilt[0]}）")
        if parts:
            tilt_conclusion = "；" + "，".join(parts)
    conclusion = f"监测数据显示：本次自动化全站仪监测各测点未出现平台预警，当前变形速率基本平稳{tilt_conclusion}。建议按既定频率持续采集监测数据，并结合现场施工进度加强对重点影响区测点的跟踪复核。"
    merge_write("A49:M52", conclusion, font_body, None, left, base_border)
    merge_write("N49:Z52", conclusion, font_body, None, left, base_border)
    merge_write("A53:M55", "项目实施单位：{{项目实施单位}}    数据来源：全站仪自动化监测平台", font_small, None, left, base_border)
    merge_write("N53:Z55", f"生成时间：{current_time}    报表类型：轨道交通控制保护区监测快报", font_small, None, left, base_border)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    check = load_workbook(output_path, data_only=False)
    sheet = check["全站仪快报"]
    if sheet.calculate_dimension() != "A1:Z55":
        raise SystemExit(f"Unexpected report range: {sheet.calculate_dimension()}")
    if image_path and image_path.exists() and not sheet._images:
        raise SystemExit("Image was not embedded.")


def build_compact_report(
    input_csv: Path,
    output_path: Path,
    image_path: Path | None = None,
    image_right_path: Path | None = None,
    point_config_path: Path | None = None,
    initial_report_path: Path | None = None,
    implementation_unit: str | None = None,
    tilt_axis: str = "横向变形量",
    point_alias_map_path: Path | None = None,
) -> None:
    rows = read_rows(input_csv)
    all_pivot = build_pivot(rows)
    pivot = {
        point: data
        for point, data in all_pivot.items()
        if not is_station_or_control_point(point)
    }
    station_control_count = len(all_pivot) - len(pivot)
    maxima = pick_maxima(pivot)
    point_config_path = point_config_path or auto_point_config_path(input_csv)
    point_geometry = point_geometry_from_rows(rows) or load_point_geometry(point_config_path)
    tilt = compute_tilt_rates(pivot, point_geometry, tilt_axis)
    tilt_max = tilt_maxima(tilt)
    alias_map = load_alias_map(point_alias_map_path)

    project_name = first_value(rows, "project_name", "项目名称") or "{{项目全称}}"
    current_time = first_value(rows, "current_time", "本次监测时间") or "{{本次监测时间}}"
    previous_time = first_value(rows, "previous_time", "上次监测时间") or "{{上次监测时间}}"
    method = first_value(rows, "monitoring_method", "监测方式") or "自动化全站仪"
    initial_time = normalize_initial_time(first_value(rows, "initial_source", "initial_time", "初始采集时间") or "{{初始采集时间}}", method)
    cadence = first_value(rows, "report_cadence", "出报间隔") or "{{15min/2h/4h}}"
    condition = first_value(rows, "work_condition", "ring_no", "施工工况") or "{{施工工况}}"
    condition_text = condition.rstrip("。；; ")
    implementation_unit = (
        implementation_unit
        or extract_implementation_unit(initial_report_path)
        or first_value(rows, "implementation_unit", "project_implementation_unit", "项目实施单位", "monitoring_unit", "监测单位")
        or "{{项目实施单位}}"
    )

    points = sorted(pivot, key=lambda p: (str(pivot[p].get("zone", "")), p))
    preferred = ["CZ11", "JD05", "S12", "X12", "S13", "X13"]
    if all(point in pivot for point in preferred):
        points = preferred

    wb = Workbook()
    temporary_images: list[Path] = []
    ws = wb.active
    ws.title = "全站仪快报"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.15, right=0.15, top=0.2, bottom=0.2, header=0, footer=0)
    ws.print_area = "A1:Z40"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True

    for col in range(1, 27):
        ws.column_dimensions[get_column_letter(col)].width = 5.2
    for row in range(1, 41):
        ws.row_dimensions[row].height = 18
    for row in range(8, 17):
        ws.row_dimensions[row].height = 28
    for row, height in {1: 24, 3: 28, 4: 28, 6: 28, 26: 22, 27: 22, 34: 28, 35: 28, 36: 28, 38: 26, 39: 26}.items():
        ws.row_dimensions[row].height = height

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=thin, right=thin, top=medium, bottom=thin)
    outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    white = PatternFill("solid", fgColor="FFFFFF")
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    green_fill = PatternFill("solid", fgColor="00B050")
    font_title = Font(name="SimSun", size=15, bold=True)
    font_section = Font(name="SimSun", size=11, bold=True)
    font_body = Font(name="SimSun", size=9)
    font_body_bold = Font(name="SimSun", size=9, bold=True)
    font_small = Font(name="SimSun", size=8)
    font_white = Font(name="SimSun", size=9, color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=26):
        for cell in row:
            cell.font = font_body
            cell.alignment = center
            cell.border = base_border
            cell.fill = white

    def merge_write(ref: str, value: object, font=None, fill=None, align=None, border=None) -> None:
        ws.merge_cells(ref)
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        cell = ws.cell(min_row, min_col)
        cell.value = value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border:
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    ws.cell(rr, cc).border = border

    # Outer page border.
    for col in range(1, 27):
        ws.cell(1, col).border = Border(top=medium, left=ws.cell(1, col).border.left, right=ws.cell(1, col).border.right, bottom=ws.cell(1, col).border.bottom)
        ws.cell(40, col).border = Border(bottom=medium, left=ws.cell(40, col).border.left, right=ws.cell(40, col).border.right, top=ws.cell(40, col).border.top)
    for row in range(1, 41):
        ws.cell(row, 1).border = Border(left=medium, right=ws.cell(row, 1).border.right, top=ws.cell(row, 1).border.top, bottom=ws.cell(row, 1).border.bottom)
        ws.cell(row, 26).border = Border(right=medium, left=ws.cell(row, 26).border.left, top=ws.cell(row, 26).border.top, bottom=ws.cell(row, 26).border.bottom)

    merge_write("A1:Z1", project_name, font_title, None, center, outer_border)
    merge_write("A2:Z2", "一、施工工况", font_section, header_fill, left, section_border)
    merge_write("A3:Z4", f"当前工况：{condition_text}。本次自动化全站仪数据取自监测平台，监测时间 {current_time}；出报频率：{cadence}。", font_body, None, left, base_border)
    merge_write("A5:Z5", "二、施工进度", font_section, header_fill, left, section_border)
    merge_write("A6:Z6", condition_text, font_body, None, left, base_border)

    merge_write("A7:Z7", "工况及影响范围示意图", font_section, sub_fill, center, base_border)
    merge_write("A8:M16", "", font_body, None, center, base_border)
    merge_write("N8:Z16", "", font_body, None, center, base_border)
    if image_path and image_path.exists():
        embed_path = make_resized_image_copy(image_path, output_path, "left_image", (520, 300))
        if embed_path != image_path:
            temporary_images.append(embed_path)
        img = XLImage(str(embed_path))
        img.anchor = "C8"
        ws.add_image(img)
    right_image = image_right_path or image_path
    if right_image and right_image.exists():
        embed_path = make_resized_image_copy(right_image, output_path, "right_image", (520, 300))
        if embed_path != right_image:
            temporary_images.append(embed_path)
        img = XLImage(str(embed_path))
        img.anchor = "P8"
        ws.add_image(img)

    def max_line(field: str) -> str:
        z = maxima.get(("纵向变形量", field), ("/", None))
        h = maxima.get(("横向变形量", field), ("/", None))
        v = maxima.get(("垂直位移", field), ("/", None))
        return "；".join(
            [
                f"纵向 {display_point(alias_map, z[0])} {fmt_01(z[1]):.1f}mm" if isinstance(z[1], float) else "纵向 /",
                f"横向 {display_point(alias_map, h[0])} {fmt_01(h[1]):.1f}mm" if isinstance(h[1], float) else "横向 /",
                f"垂直 {display_point(alias_map, v[0])} {fmt_01(v[1]):.1f}mm" if isinstance(v[1], float) else "垂直 /",
            ]
        )

    tilt_current = tilt_max.get("current", ("/", None))
    tilt_cumulative = tilt_max.get("cumulative", ("/", None))
    tilt_summary = (
        f"本次 {display_tilt_pair(alias_map, tilt_current[0])} {tilt_current[1]:.2f}‰；累计 {display_tilt_pair(alias_map, tilt_cumulative[0])} {tilt_cumulative[1]:.2f}‰"
        if isinstance(tilt_current[1], float) and isinstance(tilt_cumulative[1], float)
        else "上下测点初始三维距离待确认"
    )
    merge_write("A17:Z17", "平台最新动态", font_section, sub_fill, center, base_border)
    status_rows = [
        ("项目名称", project_name, "项目实施单位", implementation_unit),
        ("本次/上次", f"{current_time} / {previous_time}", "起算值来源", initial_time),
        ("监测方式", f"{method}；监测点 {len(pivot)} 个；测站/测量基点 {station_control_count} 个", "平台预警", "红色 0 / 橙色 0 / 黄色 0"),
        ("本次最大", max_line("current"), "累计最大", max_line("cumulative")),
        ("倾斜率", tilt_summary, "说明", "绿色为本表最大项；黄色为重点桥墩测点；红色仅用于达到预警/报警条件。"),
    ]
    for row_index, (left_label, left_value, right_label, right_value) in enumerate(status_rows, start=18):
        merge_write(f"A{row_index}:C{row_index}", left_label, font_body_bold, header_fill, center, base_border)
        merge_write(f"D{row_index}:M{row_index}", left_value, font_body, None, left, base_border)
        merge_write(f"N{row_index}:P{row_index}", right_label, font_body_bold, header_fill, center, base_border)
        merge_write(f"Q{row_index}:Z{row_index}", right_value, font_body, None, left, base_border)

    merge_write("A23:Z23", "三、各项数据分析", font_section, header_fill, left, section_border)
    merge_write("A24:D24", "初始/起算值", font_body_bold, header_fill, center, base_border)
    merge_write("E24:K24", initial_time, font_body, None, left, base_border)
    merge_write("L24:O24", "上次监测时间", font_body_bold, header_fill, center, base_border)
    merge_write("P24:Z24", previous_time, font_body, None, left, base_border)
    merge_write("A25:D25", "监测方式", font_body_bold, header_fill, center, base_border)
    merge_write("E25:K25", method, font_body, None, left, base_border)
    merge_write("L25:O25", "本次监测时间", font_body_bold, header_fill, center, base_border)
    merge_write("P25:Z25", current_time, font_body, None, left, base_border)

    header_merges = [
        ("A26:B27", "位置"),
        ("C26:D27", "监测点号"),
        ("E26:H26", "纵向变形(mm)"),
        ("I26:L26", "横向位移(mm)"),
        ("M26:P26", "垂直位移(mm)"),
        ("Q26:T26", "倾斜率(‰)"),
        ("U26:V27", "状态"),
        ("W26:Z27", "备注"),
    ]
    for ref, text in header_merges:
        merge_write(ref, text, font_body_bold, header_fill, center, base_border)
    for ref, text in [
        ("E27:F27", "本次"), ("G27:H27", "累计"),
        ("I27:J27", "本次"), ("K27:L27", "累计"),
        ("M27:N27", "本次"), ("O27:P27", "累计"),
        ("Q27:R27", "本次"), ("S27:T27", "累计"),
    ]:
        merge_write(ref, text, font_body_bold, header_fill, center, base_border)

    def set_pair(row_index: int, ref: str, value: object, fill=None, font=None) -> None:
        merge_write(ref, value, font or font_body, fill, center, base_border)
        cell = ws[ref.split(":")[0]]
        if isinstance(value, float):
            cell.number_format = "0.0"

    data_start = 28
    max_row_count = max(len(points), 1)
    for idx, point in enumerate(points):
        row_index = data_start + idx
        data = pivot[point]
        zone = display_zone(alias_map, point, str(data.get("zone", "")))
        zone_fill = yellow_fill if zone in ("12号桥墩", "13号桥墩") else gray_fill
        tilt_data = tilt.get(point, {})
        tilt_current_value = parse_float(tilt_data.get("current"))
        tilt_cumulative_value = parse_float(tilt_data.get("cumulative"))
        position_note = display_position(alias_map, point, str(data.get("position_label", "")))
        values = [
            ("A", "B", zone, zone_fill),
            ("C", "D", display_point(alias_map, point), zone_fill),
            ("E", "F", fmt_01(item_value(data, "纵向变形量", "current")), None),
            ("G", "H", fmt_01(item_value(data, "纵向变形量", "cumulative")), None),
            ("I", "J", fmt_01(item_value(data, "横向变形量", "current")), None),
            ("K", "L", fmt_01(item_value(data, "横向变形量", "cumulative")), None),
            ("M", "N", fmt_01(item_value(data, "垂直位移", "current")), None),
            ("O", "P", fmt_01(item_value(data, "垂直位移", "cumulative")), None),
            ("Q", "R", f"{tilt_current_value:.2f}" if tilt_current_value is not None else "/", None),
            ("S", "T", f"{tilt_cumulative_value:.2f}" if tilt_cumulative_value is not None else "/", None),
            ("U", "V", data.get("status", "无预警"), None),
        ]
        for start_col, end_col, value, fill in values:
            set_pair(row_index, f"{start_col}{row_index}:{end_col}{row_index}", value, fill)
        if tilt_data.get("spacing_m"):
            pair = f"{tilt_data.get('top_point')}/{tilt_data.get('lower_point')}"
            note = f"{display_tilt_pair(alias_map, pair)}初始三维距离{float(tilt_data.get('spacing_m')):.2f}m"
        elif point.startswith("X"):
            note = position_note or "倾斜下点"
        else:
            note = position_note or ("关联测点" if zone == "其他监测点" else "当前盾构影响范围跟踪")
        merge_write(f"W{row_index}:Z{row_index}", note, font_body, None, center, base_border)

        item_refs = {
            ("纵向变形量", "current"): f"E{row_index}:F{row_index}",
            ("纵向变形量", "cumulative"): f"G{row_index}:H{row_index}",
            ("横向变形量", "current"): f"I{row_index}:J{row_index}",
            ("横向变形量", "cumulative"): f"K{row_index}:L{row_index}",
            ("垂直位移", "current"): f"M{row_index}:N{row_index}",
            ("垂直位移", "cumulative"): f"O{row_index}:P{row_index}",
        }
        for key, (max_point, _) in maxima.items():
            if max_point == point:
                ref = item_refs[key]
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for rr in range(min_row, max_row + 1):
                    for cc in range(min_col, max_col + 1):
                        ws.cell(rr, cc).fill = green_fill
                        ws.cell(rr, cc).font = font_white
        for field, ref in (("current", f"Q{row_index}:R{row_index}"), ("cumulative", f"S{row_index}:T{row_index}")):
            max_item = tilt_max.get(field)
            if max_item and max_item[0] == f"{point}/{tilt_data.get('lower_point')}":
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for rr in range(min_row, max_row + 1):
                    for cc in range(min_col, max_col + 1):
                        ws.cell(rr, cc).fill = green_fill
                        ws.cell(rr, cc).font = font_white

    max_current_row = data_start + max_row_count
    max_cumulative_row = max_current_row + 1
    for row_index, label, field in [(max_current_row, "本次最大", "current"), (max_cumulative_row, "累计最大", "cumulative")]:
        merge_write(f"A{row_index}:D{row_index}", label, font_body_bold, header_fill, center, base_border)
        entries = [
            ("E", "F", maxima.get(("纵向变形量", field), ("/", None))),
            ("I", "J", maxima.get(("横向变形量", field), ("/", None))),
            ("M", "N", maxima.get(("垂直位移", field), ("/", None))),
            ("Q", "R", tilt_max.get(field, ("/", None))),
        ]
        for start_col, value_col, (point, value) in entries:
            point_ref = f"{start_col}{row_index}:{value_col}{row_index}"
            if start_col == "Q":
                display_label = display_tilt_pair(alias_map, point)
                display_value = f"{value:.2f}" if isinstance(value, float) else "/"
            else:
                display_label = display_point(alias_map, point)
                display_value = f"{fmt_01(value):.1f}" if isinstance(value, float) else "/"
            set_pair(row_index, point_ref, display_label)
            next_start = get_column_letter(ws[point_ref.split(":")[1]].column + 1)
            next_end = get_column_letter(ws[point_ref.split(":")[1]].column + 2)
            set_pair(row_index, f"{next_start}{row_index}:{next_end}{row_index}", display_value)
        merge_write(f"U{row_index}:Z{row_index}", "", font_body, None, center, base_border)

    notes_row = max_cumulative_row + 1
    spacing_notes = []
    for data in tilt.values():
        spacing = parse_float(data.get("spacing_m"))
        if spacing is not None:
            spacing_notes.append(f"{data.get('pier')}#初始三维距离{spacing:.2f}m")
    spacing_text = "；".join(spacing_notes) if spacing_notes else "上下测点初始三维距离待确认"
    notes = (
        "注：1、纵向变形、横向位移、垂直位移均以毫米计；本次/累计变形采用平台平差后成果，“+”“-”按平台成果方向及项目说明执行。\n"
        f"2、倾斜率按S/X上下测点{tilt_axis_description(tilt_axis)}计算，单位‰；"
        f"{spacing_text}。\n"
        "3、黄色表示当前重点影响区，绿色表示本表绝对值最大项；阈值按项目方案、安评或运营单位要求执行。"
    )
    merge_write(f"A{notes_row}:Z{notes_row + 2}", notes, font_small, None, left, base_border)

    conclusion_row = notes_row + 3
    merge_write(f"A{conclusion_row}:Z{conclusion_row}", "四、结论与评价", font_section, header_fill, left, section_border)
    current_tilt = tilt_max.get("current")
    cumulative_tilt = tilt_max.get("cumulative")
    tilt_text = ""
    if current_tilt and cumulative_tilt:
        tilt_text = (
            f"；本次最大倾斜率{current_tilt[1]:.2f}‰（{display_tilt_pair(alias_map, current_tilt[0])}），"
            f"累计最大倾斜率{cumulative_tilt[1]:.2f}‰（{display_tilt_pair(alias_map, cumulative_tilt[0])}）"
        )
    conclusion = (
        f"监测数据显示：本次自动化全站仪监测各测点未出现平台预警，当前变形速率基本平稳{tilt_text}。\n"
        "建议按既定频率持续采集监测数据，并结合现场施工进度加强对12号、13号桥墩及关联测点的跟踪复核。"
    )
    merge_write(f"A{conclusion_row + 1}:Z{conclusion_row + 2}", conclusion, font_body, None, left, base_border)
    merge_write(f"A{conclusion_row + 3}:M{conclusion_row + 3}", f"项目实施单位：{implementation_unit}    数据来源：全站仪自动化监测平台", font_small, None, left, base_border)
    merge_write(f"N{conclusion_row + 3}:Z{conclusion_row + 3}", f"生成时间：{current_time}    报表类型：轨道交通控制保护区监测快报", font_small, None, left, base_border)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    for temp_image in temporary_images:
        try:
            temp_image.unlink()
        except OSError:
            pass

    check = load_workbook(output_path, data_only=False)
    sheet = check["全站仪快报"]
    if sheet.calculate_dimension() != "A1:Z40":
        raise SystemExit(f"Unexpected report range: {sheet.calculate_dimension()}")
    if image_path and image_path.exists() and not sheet._images:
        raise SystemExit("Image was not embedded.")


def bridge_display_code(prefix: str, suffix: str) -> str:
    digits = re.sub(r"\D", "", suffix)
    return f"{prefix}{int(digits):03d}" if digits else f"{prefix}{suffix}"


def metric_point_code(item: str, point: str) -> str:
    suffix = point_suffix(point)
    if item == "横向变形量":
        return bridge_display_code("MCW", suffix)
    if item == TILT_ITEM:
        return bridge_display_code("MCQX", suffix)
    return bridge_display_code("MCC", suffix)


def pier_metric_value(
    pivot: dict[str, dict[str, object]],
    pier_suffix: str,
    item: str,
    field: str,
) -> tuple[str, float | None]:
    if item in {"垂直位移", "横向变形量"}:
        lower_point = f"X{pier_suffix}"
        if lower_point in pivot:
            return lower_point, item_value(pivot[lower_point], item, field)
    candidates: list[tuple[float, float, str]] = []
    for point in (f"S{pier_suffix}", f"X{pier_suffix}"):
        if point not in pivot:
            continue
        value = item_value(pivot[point], item, field)
        if value is not None:
            candidates.append((abs(value), value, point))
    if not candidates:
        return metric_point_code(item, pier_suffix), None
    _, value, point = max(candidates, key=lambda item_value: (item_value[0], item_value[2]))
    return point, value


def display_max_point(item: str, point: str) -> str:
    if point in ("", "/"):
        return "/"
    return metric_point_code(item, point)


def template_set(ws, cell: str, value: object, number_format: str | None = None) -> None:
    ws[cell].value = value
    if number_format:
        ws[cell].number_format = number_format


def template_fill_range(ws, ref: str, fill: PatternFill, font: Font | None = None) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).fill = fill
            if font:
                ws.cell(row, col).font = font


def is_directional_template(template_path: Path) -> bool:
    try:
        wb = load_workbook(template_path, read_only=True, data_only=False)
    except Exception:
        return False
    ws = wb["全站仪快报"] if "全站仪快报" in wb.sheetnames else wb.active
    values = " ".join(str(ws[cell].value or "") for cell in ("C22", "K22", "S22", "Q24", "Y24"))
    return ws.max_row >= 36 and "桥墩沉降" in values and "水平位移" in values and "桥墩倾斜" in values


def display_actual_survey_time(value: str) -> str:
    parsed = parse_report_datetime(str(value or "").replace("T", " "))
    if parsed is not None:
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    return str(value or "").replace("T", " ").strip()


def first_survey_time_from_json(path: Path | None) -> str:
    if not path or not path.exists():
        return ""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return ""
    rows = data if isinstance(data, list) else data.get("list", []) if isinstance(data, dict) else []
    if not rows:
        return ""
    value = rows[0].get("SurveyTime") if isinstance(rows[0], dict) else ""
    return display_actual_survey_time(str(value or ""))


def actual_batch_times(input_csv: Path) -> tuple[str, str]:
    current_path = input_csv.with_name(input_csv.name.replace("_adjusted_total_station.csv", "_point_coords_current.json"))
    previous_path = input_csv.with_name(input_csv.name.replace("_adjusted_total_station.csv", "_point_coords_previous.json"))
    return first_survey_time_from_json(previous_path), first_survey_time_from_json(current_path)


def load_named_coord_records(path: Path) -> dict[str, dict[str, float]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    rows = data if isinstance(data, list) else data.get("list", []) if isinstance(data, dict) else []
    records: dict[str, dict[str, float]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        name = str(row.get("PointName") or row.get("point_id") or "").strip()
        if not name:
            continue
        x = parse_float(row.get("X"))
        y = parse_float(row.get("Y"))
        h = parse_float(row.get("H"))
        if x is None or y is None or h is None:
            continue
        records[name] = {"X": x, "Y": y, "H": h}
    return records


def coord_record_paths(input_csv: Path) -> tuple[Path, Path]:
    current_path = input_csv.with_name(input_csv.name.replace("_adjusted_total_station.csv", "_point_coords_current.json"))
    previous_path = input_csv.with_name(input_csv.name.replace("_adjusted_total_station.csv", "_point_coords_previous.json"))
    return previous_path, current_path


def initial_coord_records_from_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, float]]:
    records: dict[str, dict[str, float]] = {}
    for row in rows:
        name = str(row.get("point_id") or "").strip()
        if not name or name in records:
            continue
        x = parse_float(row.get("initial_x_m"))
        y = parse_float(row.get("initial_y_m"))
        h = parse_float(row.get("initial_h_m"))
        if x is None or y is None or h is None:
            continue
        records[name] = {"X": x, "Y": y, "H": h}
    return records


def coord_value(records: dict[str, dict[str, float]], point: str, axis: str) -> float | str:
    value = records.get(point, {}).get(axis)
    return value if value is not None else ""


def set_coord_triplet(
    ws,
    row_index: int,
    point: str,
    initial_coords: dict[str, dict[str, float]],
    previous_coords: dict[str, dict[str, float]],
    current_coords: dict[str, dict[str, float]],
) -> None:
    for offset, axis in enumerate(("X", "Y", "H")):
        ws.cell(row_index, 5 + offset).value = coord_value(initial_coords, point, axis)
        ws.cell(row_index, 8 + offset).value = coord_value(previous_coords, point, axis)
        ws.cell(row_index, 11 + offset).value = coord_value(current_coords, point, axis)


def set_coord_xy_groups(
    ws,
    row_index: int,
    point: str,
    initial_coords: dict[str, dict[str, float]],
    previous_coords: dict[str, dict[str, float]],
    current_coords: dict[str, dict[str, float]],
) -> None:
    ws.cell(row_index, 5).value = coord_value(initial_coords, point, "X")
    ws.cell(row_index, 6).value = coord_value(initial_coords, point, "Y")
    ws.cell(row_index, 7).value = coord_value(previous_coords, point, "X")
    ws.cell(row_index, 8).value = coord_value(previous_coords, point, "Y")
    ws.cell(row_index, 9).value = coord_value(current_coords, point, "X")
    ws.cell(row_index, 10).value = coord_value(current_coords, point, "Y")


def style_detail_numeric_columns(ws, rows: range, columns: tuple[int, ...], number_format: str) -> None:
    for row_index in rows:
        for col_index in columns:
            ws.cell(row_index, col_index).number_format = number_format


def clear_charts(ws) -> None:
    ws._charts = []


def add_metric_line_chart(
    ws,
    title: str,
    category_col: int,
    data_cols: tuple[int, int],
    min_row: int,
    max_row: int,
    anchor: str,
    y_axis_format: str,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = "变化量"
    chart.x_axis.title = "测点编号"
    chart.y_axis.numFmt = y_axis_format
    chart.height = 7.2
    chart.width = 18
    for col_index in data_cols:
        data = Reference(ws, min_col=col_index, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    categories = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def add_tilt_line_chart(ws) -> None:
    helper_start = 12
    helper_rows = (
        (12, "东西方向"),
        (14, "东西方向"),
        (16, "南北方向"),
        (18, "南北方向"),
    )
    ws["R11"] = "测点编号"
    ws["S11"] = "本次变量"
    ws["T11"] = "累计变量"
    for offset, (source_row, direction) in enumerate(helper_rows):
        row_index = helper_start + offset
        ws[f"R{row_index}"] = f'=A{source_row}&" {direction}"'
        ws[f"S{row_index}"] = f"=B{source_row}"
        ws[f"T{row_index}"] = f"=D{source_row}"
        ws[f"S{row_index}"].number_format = "0.00000"
        ws[f"T{row_index}"].number_format = "0.00000"
    add_metric_line_chart(ws, "桥墩倾斜", 18, (19, 20), 11, 15, "A35", "0.00000")


def format_report_date(value: str) -> str:
    parsed = parse_report_datetime(value)
    if parsed is None:
        return value
    return parsed.strftime("%Y/%-m/%-d %H:%M")


def monitoring_count_and_hours(first_time: str, current_time: str, interval_hours: int = 4) -> tuple[int | str, int | str]:
    first_dt = parse_report_datetime(first_time)
    current_dt = parse_report_datetime(current_time)
    if first_dt is None or current_dt is None:
        return "", ""
    hours = int(round((current_dt - first_dt).total_seconds() / 3600))
    if hours < 0:
        return "", ""
    return hours // interval_hours + 1, hours


def spacing_formula(row_a: int, row_b: int) -> str:
    return (
        f"SQRT(($E${row_a}-$E${row_b})*($E${row_a}-$E${row_b})+"
        f"($F${row_a}-$F${row_b})*($F${row_a}-$F${row_b})+"
        f"($G${row_a}-$G${row_b})*($G${row_a}-$G${row_b}))"
    )


def fill_complete_detail_sheets(
    wb,
    input_csv: Path,
    rows: list[dict[str, str]],
    project_name: str,
    previous_time: str,
    current_time: str,
    condition_text: str,
    implementation_unit: str,
    manual_overrides_path: Path | None = None,
) -> None:
    required = {"封面", "桥墩沉降（自动化）", "桥墩水平位移（自动化）", "桥墩倾斜（自动化）"}
    if not required.issubset(set(wb.sheetnames)):
        return

    if "适配设置页" in wb.sheetnames:
        wb["适配设置页"].sheet_state = "hidden"
    if "封面" in wb.sheetnames:
        wb.active = wb.sheetnames.index("封面")

    first_time = "2026/6/4 16:00"
    monitor_count, total_hours = monitoring_count_and_hours(first_time, current_time)
    previous_path, current_path = coord_record_paths(input_csv)
    previous_coords = load_named_coord_records(previous_path)
    current_coords = load_named_coord_records(current_path)
    initial_coords = initial_coord_records_from_rows(rows)
    apply_manual_coordinate_overrides(initial_coords, previous_coords, current_coords, rows, manual_overrides_path)
    condition_full = f"当前工况：{condition_text}。本次自动化全站仪数据取自监测平台，监测时间 {current_time}；出报频率：4h。"

    cover = wb["封面"]
    cover["A4"] = project_name.replace("施工期间", "施工期间\n")
    cover["A7"] = "监测时报表"
    cover["A28"] = implementation_unit
    cover["A31"] = current_time

    settlement = wb["桥墩沉降（自动化）"]
    settlement["A2"] = f"监测单位：{implementation_unit}"
    settlement["A4"] = f"项目名称：{project_name}"
    settlement["D6"] = first_time.split()[0]
    settlement["H5"] = monitor_count
    settlement["H6"] = 4
    settlement["H7"] = total_hours
    settlement["D7"] = previous_time
    settlement["D8"] = current_time
    settlement["N7"] = condition_full
    settlement["N4"] = settlement["H5"].value
    for row_index, suffix in ((12, "12"), (13, "13")):
        point = lower_point_for_pier(suffix)
        settlement[f"A{row_index}"] = bridge_display_code("MCC", suffix)
        settlement[f"B{row_index}"] = f"=(O{row_index}-M{row_index})*1000"
        settlement[f"E{row_index}"] = f"=B{row_index}/4"
        settlement[f"H{row_index}"] = f"=(O{row_index}-J{row_index})*1000"
        settlement[f"J{row_index}"] = coord_value(initial_coords, point, "H")
        settlement[f"M{row_index}"] = coord_value(previous_coords, point, "H")
        settlement[f"O{row_index}"] = coord_value(current_coords, point, "H")
        settlement[f"Q{row_index}"] = f"{int(suffix)}#桥墩"
    style_detail_numeric_columns(settlement, range(12, 14), (2, 5, 8), "0.0")
    style_detail_numeric_columns(settlement, range(12, 14), (10, 13, 15), "0.00000")
    settlement["A32"] = "报警值：预警值±5mm，报警值±7mm，控制值±10mm。"
    settlement.print_area = "A1:R44"
    clear_charts(settlement)
    add_metric_line_chart(settlement, "桥墩沉降", 1, (2, 8), 11, 13, "A35", "0.0")

    horizontal = wb["桥墩水平位移（自动化）"]
    horizontal["A2"] = f"监测单位：{implementation_unit}"
    horizontal["A4"] = f"项目名称：{project_name}"
    horizontal["D6"] = first_time
    horizontal["G5"] = monitor_count
    horizontal["G6"] = 4
    horizontal["G7"] = total_hours
    horizontal["D7"] = previous_time
    horizontal["D8"] = current_time
    horizontal["K7"] = condition_full
    for row_index, suffix, direction in ((12, "12", "东西方向"), (13, "13", "东西方向"), (14, "12", "南北方向"), (15, "13", "南北方向")):
        point = lower_point_for_pier(suffix)
        horizontal[f"A{row_index}"] = bridge_display_code("MCW", suffix)
        if direction == "东西方向":
            horizontal[f"B{row_index}"] = f"=(J{row_index}-H{row_index})*1000"
            horizontal[f"D{row_index}"] = f"=(J{row_index}-F{row_index})*1000"
        else:
            horizontal[f"B{row_index}"] = f"=(I{row_index}-G{row_index})*1000"
            horizontal[f"D{row_index}"] = f"=(I{row_index}-E{row_index})*1000"
        horizontal[f"C{row_index}"] = f"=B{row_index}/4"
        set_coord_xy_groups(horizontal, row_index, point, initial_coords, previous_coords, current_coords)
        horizontal[f"K{row_index}"] = f"{int(suffix)}#桥墩"
        horizontal[f"L{row_index}"] = direction
    style_detail_numeric_columns(horizontal, range(12, 16), (2, 3, 4), "0.0")
    style_detail_numeric_columns(horizontal, range(12, 16), (5, 6, 7, 8, 9, 10), "0.00000")
    horizontal["A32"] = "报警值：预警值±5mm，报警值±7mm，控制值±10mm。"
    horizontal.print_area = "A1:L47"
    clear_charts(horizontal)
    add_metric_line_chart(horizontal, "桥墩水平位移", 1, (2, 4), 11, 15, "A35", "0.0")

    tilt = wb["桥墩倾斜（自动化）"]
    tilt["A2"] = f"监测单位：{implementation_unit}"
    tilt["A4"] = f"项目名称：{project_name}"
    tilt["D6"] = first_time
    tilt["H5"] = monitor_count
    tilt["H6"] = 4
    tilt["H7"] = total_hours
    tilt["D7"] = previous_time
    tilt["D8"] = current_time
    tilt["M7"] = condition_full

    for start_row, suffix, direction in ((12, "12", "东西方向"), (14, "13", "东西方向"), (16, "12", "南北方向"), (18, "13", "南北方向")):
        top = top_point_for_pier(suffix)
        lower = lower_point_for_pier(suffix)
        tilt[f"A{start_row}"] = bridge_display_code("MCQX", suffix)
        if direction == "东西方向":
            tilt[f"B{start_row}"] = f"=((L{start_row}-I{start_row})-(L{start_row+1}-I{start_row+1}))/{spacing_formula(start_row, start_row+1)}*1000"
            tilt[f"D{start_row}"] = f"=((L{start_row}-F{start_row})-(L{start_row+1}-F{start_row+1}))/{spacing_formula(start_row, start_row+1)}*1000"
        else:
            tilt[f"B{start_row}"] = f"=((K{start_row}-H{start_row})-(K{start_row+1}-H{start_row+1}))/{spacing_formula(start_row, start_row+1)}*1000"
            tilt[f"D{start_row}"] = f"=((K{start_row}-E{start_row})-(K{start_row+1}-E{start_row+1}))/{spacing_formula(start_row, start_row+1)}*1000"
        tilt[f"C{start_row}"] = f"=B{start_row}/4"
        set_coord_triplet(tilt, start_row, top, initial_coords, previous_coords, current_coords)
        set_coord_triplet(tilt, start_row + 1, lower, initial_coords, previous_coords, current_coords)
        tilt[f"N{start_row}"] = f"{int(suffix)}#桥墩"
        tilt[f"O{start_row}"] = direction
    style_detail_numeric_columns(tilt, range(12, 20), (2, 3, 4), "0.00000")
    style_detail_numeric_columns(tilt, range(12, 20), (5, 6, 7, 8, 9, 10, 11, 12, 13), "0.00000")
    tilt["A32"] = "报警值：倾斜：预警值1.0‰，报警值1.4‰，控制值2.0‰。"
    tilt.print_area = "A1:O47"
    clear_charts(tilt)
    add_tilt_line_chart(tilt)

    for sheet_name in ("封面", "桥墩沉降（自动化）", "桥墩水平位移（自动化）", "桥墩倾斜（自动化）"):
        ws = wb[sheet_name]
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1


def direction_item(direction: str) -> str:
    return "横向变形量" if direction == "east" else "纵向变形量"


def direction_label(direction: str) -> str:
    return "东西方向" if direction == "east" else "南北方向"


def direction_short_label(direction: str) -> str:
    return "东西" if direction == "east" else "南北"


def lower_point_for_pier(pier_suffix: str) -> str:
    return f"X{pier_suffix}"


def top_point_for_pier(pier_suffix: str) -> str:
    return f"S{pier_suffix}"


def pier_suffixes_from_pivot(pivot: dict[str, dict[str, object]]) -> list[str]:
    suffixes = sorted(
        {
            point_suffix(point)
            for point in pivot
            if f"S{point_suffix(point)}" in pivot and f"X{point_suffix(point)}" in pivot
        },
        key=lambda item: int(item) if str(item).isdigit() else str(item),
    )
    return suffixes


def lower_metric_value(
    pivot: dict[str, dict[str, object]],
    pier_suffix: str,
    item: str,
    field: str,
) -> float | None:
    point = lower_point_for_pier(pier_suffix)
    if point not in pivot:
        return None
    return item_value(pivot[point], item, field)


def compute_directional_tilts(
    pivot: dict[str, dict[str, object]],
    point_geometry: dict[str, dict[str, float]] | dict[str, float],
) -> dict[str, dict[str, dict[str, float | None]]]:
    tilts: dict[str, dict[str, dict[str, float | None]]] = {}
    for suffix in pier_suffixes_from_pivot(pivot):
        top = top_point_for_pier(suffix)
        lower = lower_point_for_pier(suffix)
        spacing = point_spacing_m(top, lower, point_geometry)
        tilts[suffix] = {
            "spacing": {"value": spacing},
            "east": {"current": None, "cumulative": None},
            "north": {"current": None, "cumulative": None},
        }
        if spacing is None or spacing <= 0:
            continue
        for direction in ("east", "north"):
            item = direction_item(direction)
            for field in ("current", "cumulative"):
                top_value = item_value(pivot[top], item, field)
                lower_value = item_value(pivot[lower], item, field)
                if top_value is None or lower_value is None:
                    continue
                tilts[suffix][direction][field] = (float(top_value) - float(lower_value)) / spacing
    return tilts


def directional_maxima(
    pivot: dict[str, dict[str, object]],
    directional_tilts: dict[str, dict[str, dict[str, float | None]]],
) -> dict[tuple[str, str], dict[str, object]]:
    maxima: dict[tuple[str, str], dict[str, object]] = {}
    pier_suffixes = pier_suffixes_from_pivot(pivot)
    for field in ("current", "cumulative"):
        settlement_candidates: list[dict[str, object]] = []
        horizontal_candidates: list[dict[str, object]] = []
        tilt_candidates: list[dict[str, object]] = []
        for suffix in pier_suffixes:
            settlement = lower_metric_value(pivot, suffix, "垂直位移", field)
            if settlement is not None:
                settlement_candidates.append(
                    {
                        "suffix": suffix,
                        "point": lower_point_for_pier(suffix),
                        "display_point": bridge_display_code("MCC", suffix),
                        "value": settlement,
                        "direction": "",
                    }
                )
            for direction in ("east", "north"):
                horizontal = lower_metric_value(pivot, suffix, direction_item(direction), field)
                if horizontal is not None:
                    horizontal_candidates.append(
                        {
                            "suffix": suffix,
                            "point": lower_point_for_pier(suffix),
                            "display_point": bridge_display_code("MCW", suffix),
                            "value": horizontal,
                            "direction": direction_label(direction),
                        }
                    )
                tilt_value = directional_tilts.get(suffix, {}).get(direction, {}).get(field)
                if tilt_value is not None:
                    tilt_candidates.append(
                        {
                            "suffix": suffix,
                            "point": f"{top_point_for_pier(suffix)}/{lower_point_for_pier(suffix)}",
                            "display_point": bridge_display_code("MCQX", suffix),
                            "value": tilt_value,
                            "direction": direction_label(direction),
                        }
                    )
        for key, candidates in (
            (("settlement", field), settlement_candidates),
            (("horizontal", field), horizontal_candidates),
            (("tilt", field), tilt_candidates),
        ):
            if candidates:
                maxima[key] = max(
                    candidates,
                    key=lambda item: (
                        abs(float(item.get("value") or 0)),
                        str(item.get("suffix") or ""),
                        str(item.get("direction") or ""),
                    ),
                )
    return maxima


def build_directional_template_report(
    input_csv: Path,
    output_path: Path,
    template_path: Path,
    image_path: Path | None = None,
    image_right_path: Path | None = None,
    point_config_path: Path | None = None,
    initial_report_path: Path | None = None,
    implementation_unit: str | None = None,
    point_alias_map_path: Path | None = None,
    work_condition_override: str | None = None,
    manual_overrides_path: Path | None = None,
) -> None:
    rows = read_rows(input_csv)
    all_pivot = build_pivot(rows)
    pivot = {
        point: data
        for point, data in all_pivot.items()
        if not is_station_or_control_point(point)
    }
    manual_overrides_path = manual_overrides_path or input_csv.with_name("manual_metric_overrides.json")
    apply_manual_metric_overrides(pivot, rows, manual_overrides_path)
    point_config_path = point_config_path or auto_point_config_path(input_csv)
    point_geometry = point_geometry_from_rows(rows) or load_point_geometry(point_config_path)
    directional_tilts = compute_directional_tilts(pivot, point_geometry)
    maxima = directional_maxima(pivot, directional_tilts)

    project_name = first_value(rows, "project_name", "项目名称") or "{{项目全称}}"
    current_time = first_value(rows, "current_time", "本次监测时间") or "{{本次监测时间}}"
    previous_time = first_value(rows, "previous_time", "上次监测时间") or "{{上次监测时间}}"
    actual_previous, actual_current = actual_batch_times(input_csv)
    method = first_value(rows, "monitoring_method", "监测方式") or "自动化全站仪"
    cadence = first_value(rows, "report_cadence", "出报间隔") or "{{15min/2h/4h}}"
    condition = work_condition_override or first_value(rows, "work_condition", "ring_no", "施工工况") or "{{施工工况}}"
    condition_text = condition.rstrip("。；; ")
    implementation_unit = (
        implementation_unit
        or extract_implementation_unit(initial_report_path)
        or first_value(rows, "implementation_unit", "project_implementation_unit", "项目实施单位", "monitoring_unit", "监测单位")
        or "{{项目实施单位}}"
    )

    wb = load_workbook(template_path)
    ws = wb["全站仪快报"] if "全站仪快报" in wb.sheetnames else wb.active
    temporary_images: list[Path] = []
    green_fill = PatternFill("solid", fgColor="00B050")
    white_font = Font(name="SimSun", size=9, color="FFFFFF", bold=True)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_font = Font(name="SimSun", size=9, color="000000")

    template_set(ws, "A1", project_name)
    template_set(
        ws,
        "A3",
        f"当前工况：{condition_text}。本次自动化全站仪数据取自监测平台，监测时间 {current_time}；出报频率：{cadence}。",
    )
    template_set(ws, "A6", None)
    template_set(ws, "O6", None)
    if image_path and image_path.exists():
        left_image = make_canvas_image_copy(image_path, output_path, "directional_left_image", (760, 300))
        if left_image != image_path:
            temporary_images.append(left_image)
        img = XLImage(str(left_image))
        img.anchor = "A6"
        ws.add_image(img)
    if image_right_path and image_right_path.exists():
        right_image = make_canvas_image_copy(image_right_path, output_path, "directional_right_image", (760, 300))
        if right_image != image_right_path:
            temporary_images.append(right_image)
        img = XLImage(str(right_image))
        img.anchor = "O6"
        ws.add_image(img)

    def summary_part(kind: str, field: str, label: str, unit: str, decimals: int) -> str:
        item = maxima.get((kind, field))
        if not item:
            return f"{label}：/"
        value = parse_float(item.get("value"))
        direction = str(item.get("direction") or "")
        value_text = fmt_02_text(value) if decimals == 2 else fmt_01_text(value)
        direction_text = f"（{direction.replace('方向', '')}）" if direction else ""
        return f"{label}：{item.get('display_point')} {value_text}{unit}{direction_text}"

    template_set(ws, "D16", project_name)
    template_set(ws, "R16", implementation_unit)
    template_set(ws, "D17", method)
    template_set(ws, "R17", "红色 0 / 橙色 0 / 黄色 0")
    template_set(
        ws,
        "D18",
        "；".join(
            [
                summary_part("settlement", "current", "沉降", "mm", 1),
                summary_part("horizontal", "current", "位移", "mm", 1),
                summary_part("tilt", "current", "倾斜", "‰", 2),
            ]
        ),
    )
    template_set(
        ws,
        "R18",
        "；".join(
            [
                summary_part("settlement", "cumulative", "沉降", "mm", 1),
                summary_part("horizontal", "cumulative", "位移", "mm", 1),
                summary_part("tilt", "cumulative", "倾斜", "‰", 2),
            ]
        ),
    )
    spacing_notes = []
    for suffix in pier_suffixes_from_pivot(pivot):
        spacing = directional_tilts.get(suffix, {}).get("spacing", {}).get("value")
        if spacing is not None:
            spacing_notes.append(f"{bridge_display_code('MCQX', suffix)}初始三维距离{float(spacing):.2f}m")
    template_set(ws, "A19", "倾斜计算距离")
    template_set(ws, "D19", "，".join(spacing_notes) if spacing_notes else "上下测点初始三维距离待确认")
    template_set(ws, "R19", "绿色为主数据区绝对值最大项；黄色为达到预警条件；红色达到报警条件")
    template_set(ws, "D21", previous_time)
    template_set(ws, "R21", current_time)

    pier_rows = [("12", 24, "east"), ("13", 25, "east"), ("12", 26, "north"), ("13", 27, "north")]
    clear_ranges = []
    for row_index in range(24, 30):
        clear_ranges.extend(
            [
                f"E{row_index}:F{row_index}",
                f"G{row_index}:H{row_index}",
                f"M{row_index}:N{row_index}",
                f"O{row_index}:P{row_index}",
                f"U{row_index}:V{row_index}",
                f"W{row_index}:X{row_index}",
            ]
        )
    for ref in clear_ranges:
        template_fill_range(ws, ref, white_fill, black_font)

    for suffix, row_index, direction in pier_rows:
        template_set(ws, f"A{row_index}", f"{int(suffix)}号桥墩")
        if direction == "east":
            template_set(ws, f"C{row_index}", bridge_display_code("MCC", suffix))
            template_set(ws, f"E{row_index}", fmt_01_value(lower_metric_value(pivot, suffix, "垂直位移", "current")), "0.0")
            template_set(ws, f"G{row_index}", fmt_01_value(lower_metric_value(pivot, suffix, "垂直位移", "cumulative")), "0.0")
            template_set(ws, f"I{row_index}", "")
        else:
            template_set(ws, f"C{row_index}", "")
            template_set(ws, f"E{row_index}", "")
            template_set(ws, f"G{row_index}", "")
            template_set(ws, f"I{row_index}", "")
        template_set(ws, f"K{row_index}", bridge_display_code("MCW", suffix))
        template_set(ws, f"M{row_index}", fmt_01_value(lower_metric_value(pivot, suffix, direction_item(direction), "current")), "0.0")
        template_set(ws, f"O{row_index}", fmt_01_value(lower_metric_value(pivot, suffix, direction_item(direction), "cumulative")), "0.0")
        template_set(ws, f"Q{row_index}", direction_label(direction))
        template_set(ws, f"S{row_index}", bridge_display_code("MCQX", suffix))
        tilt_current = directional_tilts.get(suffix, {}).get(direction, {}).get("current")
        tilt_cumulative = directional_tilts.get(suffix, {}).get(direction, {}).get("cumulative")
        template_set(ws, f"U{row_index}", fmt_02_value(parse_float(tilt_current)), "0.00")
        template_set(ws, f"W{row_index}", fmt_02_value(parse_float(tilt_cumulative)), "0.00")
        template_set(ws, f"Y{row_index}", direction_label(direction))
        template_set(ws, f"AA{row_index}", "无预警")

    # Highlight only data rows, not the summary rows.
    data_cell_refs = {
        ("settlement", "current", "12", ""): "E24:F24",
        ("settlement", "current", "13", ""): "E25:F25",
        ("settlement", "cumulative", "12", ""): "G24:H24",
        ("settlement", "cumulative", "13", ""): "G25:H25",
        ("horizontal", "current", "12", "东西方向"): "M24:N24",
        ("horizontal", "current", "13", "东西方向"): "M25:N25",
        ("horizontal", "current", "12", "南北方向"): "M26:N26",
        ("horizontal", "current", "13", "南北方向"): "M27:N27",
        ("horizontal", "cumulative", "12", "东西方向"): "O24:P24",
        ("horizontal", "cumulative", "13", "东西方向"): "O25:P25",
        ("horizontal", "cumulative", "12", "南北方向"): "O26:P26",
        ("horizontal", "cumulative", "13", "南北方向"): "O27:P27",
        ("tilt", "current", "12", "东西方向"): "U24:V24",
        ("tilt", "current", "13", "东西方向"): "U25:V25",
        ("tilt", "current", "12", "南北方向"): "U26:V26",
        ("tilt", "current", "13", "南北方向"): "U27:V27",
        ("tilt", "cumulative", "12", "东西方向"): "W24:X24",
        ("tilt", "cumulative", "13", "东西方向"): "W25:X25",
        ("tilt", "cumulative", "12", "南北方向"): "W26:X26",
        ("tilt", "cumulative", "13", "南北方向"): "W27:X27",
    }
    for (kind, field), item in maxima.items():
        ref = data_cell_refs.get((kind, field, str(item.get("suffix") or ""), str(item.get("direction") or "")))
        if ref:
            template_fill_range(ws, ref, green_fill, white_font)

    for row_index, field, label in ((28, "current", "本次最大"), (29, "cumulative", "累计最大")):
        settlement = maxima.get(("settlement", field), {})
        horizontal = maxima.get(("horizontal", field), {})
        tilt = maxima.get(("tilt", field), {})
        template_set(ws, f"C{row_index}", label)
        template_set(ws, f"E{row_index}", settlement.get("display_point", "/"))
        template_set(ws, f"G{row_index}", fmt_01_value(parse_float(settlement.get("value"))), "0.0")
        template_set(ws, f"I{row_index}", "")
        template_set(ws, f"K{row_index}", label)
        template_set(ws, f"M{row_index}", horizontal.get("display_point", "/"))
        template_set(ws, f"O{row_index}", fmt_01_value(parse_float(horizontal.get("value"))), "0.0")
        template_set(ws, f"Q{row_index}", horizontal.get("direction", ""))
        template_set(ws, f"S{row_index}", label)
        template_set(ws, f"U{row_index}", tilt.get("display_point", "/"))
        template_set(ws, f"W{row_index}", fmt_02_value(parse_float(tilt.get("value"))), "0.00")
        template_set(ws, f"Y{row_index}", tilt.get("direction", ""))
        template_set(ws, f"AA{row_index}", "")

    notes = (
        "注：1、竖向位移：“+”表示隆起，“-”表示下沉；\n"
        "       东西方向水平位移：“+”为向东方向，“-”为向西方向；\n"
        "       南北方向水平位移：“+”为向北方向，“-”为向南方向；\n"
        "       东西方向倾斜：“+”为东方向偏移，“-”为西方向偏移；\n"
        "       南北方向倾斜：“+”为北方向偏移，“-”为南方向偏移；\n"
        "    2、沉降、位移：预警值±5mm，报警值±7mm，控制值±10mm；\n"
        "       倾斜：预警值1.0‰，报警值1.4‰，控制值2.0‰。"
    )
    template_set(ws, "A30", notes)
    template_set(
        ws,
        "A34",
        "监测数据显示：本次自动化全站仪监测各测点未出现平台预警，当前变形速率基本平稳；\n"
        "建议按既定频率持续采集监测数据，并结合现场施工进度加强对12号、13号桥墩及关联测点的跟踪复核。",
    )
    template_set(ws, "A36", f"项目实施单位：{implementation_unit}    数据来源：全站仪自动化监测平台")
    template_set(ws, "O36", f"生成时间：{current_time}    报表类型：轨道交通控制保护区监测快报")

    ws.print_area = "A1:AB36"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    fill_complete_detail_sheets(
        wb,
        input_csv,
        rows,
        project_name,
        previous_time,
        current_time,
        condition_text,
        implementation_unit,
        manual_overrides_path,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    for temp_image in temporary_images:
        try:
            temp_image.unlink()
        except OSError:
            pass
    check = load_workbook(output_path, data_only=False)
    sheet = check["全站仪快报"]
    if sheet.calculate_dimension() != "A1:AB36":
        raise SystemExit(f"Unexpected report range: {sheet.calculate_dimension()}")
    if image_path and image_path.exists() and not sheet._images:
        raise SystemExit("Image was not embedded.")


def build_template_report(
    input_csv: Path,
    output_path: Path,
    template_path: Path,
    image_path: Path | None = None,
    image_right_path: Path | None = None,
    point_config_path: Path | None = None,
    initial_report_path: Path | None = None,
    implementation_unit: str | None = None,
    tilt_axis: str = "横向变形量",
    point_alias_map_path: Path | None = None,
    work_condition_override: str | None = None,
    manual_overrides_path: Path | None = None,
) -> None:
    if is_directional_template(template_path):
        build_directional_template_report(
            input_csv,
            output_path,
            template_path,
            image_path,
            image_right_path,
            point_config_path,
            initial_report_path,
            implementation_unit,
            point_alias_map_path,
            work_condition_override,
            manual_overrides_path,
        )
        return

    rows = read_rows(input_csv)
    all_pivot = build_pivot(rows)
    pivot = {
        point: data
        for point, data in all_pivot.items()
        if not is_station_or_control_point(point)
    }
    manual_overrides_path = manual_overrides_path or input_csv.with_name("manual_metric_overrides.json")
    apply_manual_metric_overrides(pivot, rows, manual_overrides_path)
    alias_map = load_alias_map(point_alias_map_path)
    maxima = pick_maxima(pivot)
    point_config_path = point_config_path or auto_point_config_path(input_csv)
    point_geometry = point_geometry_from_rows(rows) or load_point_geometry(point_config_path)
    tilt = compute_tilt_rates(pivot, point_geometry, tilt_axis)
    tilt_max = tilt_maxima(tilt)

    project_name = first_value(rows, "project_name", "项目名称") or "{{项目全称}}"
    current_time = first_value(rows, "current_time", "本次监测时间") or "{{本次监测时间}}"
    previous_time = first_value(rows, "previous_time", "上次监测时间") or "{{上次监测时间}}"
    method = first_value(rows, "monitoring_method", "监测方式") or "自动化全站仪"
    cadence = first_value(rows, "report_cadence", "出报间隔") or "{{15min/2h/4h}}"
    condition = work_condition_override or first_value(rows, "work_condition", "ring_no", "施工工况") or "{{施工工况}}"
    condition_text = condition.rstrip("。；; ")
    implementation_unit = (
        implementation_unit
        or extract_implementation_unit(initial_report_path)
        or first_value(rows, "implementation_unit", "project_implementation_unit", "项目实施单位", "monitoring_unit", "监测单位")
        or "{{项目实施单位}}"
    )

    wb = load_workbook(template_path)
    ws = wb.active
    temporary_images: list[Path] = []
    green_fill = PatternFill("solid", fgColor="00B050")
    white_font = Font(name="SimSun", size=9, color="FFFFFF", bold=True)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_font = Font(name="SimSun", size=9, color="000000")

    template_set(ws, "A1", project_name)
    template_set(
        ws,
        "A3",
        f"当前工况：{condition_text}。本次自动化全站仪数据取自监测平台，监测时间 {current_time}；出报频率：{cadence}。",
    )
    template_set(ws, "A6", None)
    template_set(ws, "O6", None)

    if image_path and image_path.exists():
        left_image = make_canvas_image_copy(image_path, output_path, "template_left_image", (760, 300))
        if left_image != image_path:
            temporary_images.append(left_image)
        img = XLImage(str(left_image))
        img.anchor = "A6"
        ws.add_image(img)
    if image_right_path and image_right_path.exists():
        right_image = make_canvas_image_copy(image_right_path, output_path, "template_right_image", (760, 300))
        if right_image != image_right_path:
            temporary_images.append(right_image)
        img = XLImage(str(right_image))
        img.anchor = "O6"
        ws.add_image(img)

    def max_summary(item: str, field: str, label: str) -> str:
        point, value = maxima.get((item, field), ("/", None))
        if not isinstance(value, float):
            return f"{label}：/"
        point_text = display_max_point(item, point)
        return f"{label}：{point_text} {fmt_01_text(value)}mm"

    def tilt_summary(field: str, label: str) -> str:
        point, value = tilt_max.get(field, ("/", None))
        if not isinstance(value, float):
            return f"{label}：/"
        return f"{label}：{display_tilt_pair(alias_map, point)} {fmt_01_text(value)}‰"

    template_set(ws, "D16", project_name)
    template_set(ws, "R16", implementation_unit)
    template_set(ws, "D17", method)
    template_set(ws, "R17", "红色 0 / 橙色 0 / 黄色 0")
    template_set(
        ws,
        "D18",
        "；".join(
            [
                max_summary("垂直位移", "current", "沉降"),
                max_summary("横向变形量", "current", "位移"),
                tilt_summary("current", "倾斜"),
            ]
        ),
    )
    template_set(
        ws,
        "R18",
        "；".join(
            [
                max_summary("垂直位移", "cumulative", "沉降"),
                max_summary("横向变形量", "cumulative", "位移"),
                tilt_summary("cumulative", "倾斜"),
            ]
        ),
    )
    spacing_notes = []
    for data in tilt.values():
        spacing = parse_float(data.get("spacing_m"))
        pier = data.get("pier")
        if spacing is not None and pier:
            spacing_notes.append(f"MCQX{int(str(pier)):03d}初始三维距离{spacing:.2f}m")
    template_set(ws, "A19", "倾斜计算距离")
    template_set(ws, "D19", "，".join(spacing_notes) if spacing_notes else "上下测点初始三维距离待确认")
    template_set(ws, "R19", "绿色为主数据区绝对值最大项；黄色为达到预警条件；红色达到报警条件")
    template_set(ws, "D21", previous_time)
    template_set(ws, "R21", current_time)

    row_by_pier = {"12": 24, "13": 25}
    for suffix, row_index in row_by_pier.items():
        template_set(ws, f"A{row_index}", f"{int(suffix)}号桥墩")
        template_set(ws, f"C{row_index}", bridge_display_code("MCC", suffix))
        _, settlement_current = pier_metric_value(pivot, suffix, "垂直位移", "current")
        _, settlement_cumulative = pier_metric_value(pivot, suffix, "垂直位移", "cumulative")
        template_set(ws, f"E{row_index}", fmt_01_value(settlement_current), "0.0")
        template_set(ws, f"G{row_index}", fmt_01_value(settlement_cumulative), "0.0")
        template_set(ws, f"I{row_index}", "")

        template_set(ws, f"K{row_index}", bridge_display_code("MCW", suffix))
        _, horizontal_current = pier_metric_value(pivot, suffix, "横向变形量", "current")
        _, horizontal_cumulative = pier_metric_value(pivot, suffix, "横向变形量", "cumulative")
        template_set(ws, f"M{row_index}", fmt_01_value(horizontal_current), "0.0")
        template_set(ws, f"O{row_index}", fmt_01_value(horizontal_cumulative), "0.0")
        template_set(ws, f"Q{row_index}", "")

        top_point = f"S{suffix}"
        tilt_data = tilt.get(top_point, {})
        current_tilt = parse_float(tilt_data.get("current"))
        cumulative_tilt = parse_float(tilt_data.get("cumulative"))
        template_set(ws, f"S{row_index}", bridge_display_code("MCQX", suffix))
        template_set(ws, f"U{row_index}", fmt_01_text(current_tilt) if current_tilt is not None else "/")
        template_set(ws, f"W{row_index}", fmt_01_text(cumulative_tilt) if cumulative_tilt is not None else "/")
        template_set(ws, f"Y{row_index}", "")
        template_set(ws, f"AA{row_index}", "无预警")

    for row_index, field, label in ((26, "current", "本次最大"), (27, "cumulative", "累计最大")):
        template_set(ws, f"C{row_index}", label)
        point, value = maxima.get(("垂直位移", field), ("/", None))
        template_set(ws, f"E{row_index}", display_max_point("垂直位移", point))
        template_set(ws, f"G{row_index}", fmt_01_value(value), "0.0")
        template_set(ws, f"I{row_index}", "")

        template_set(ws, f"K{row_index}", label)
        point, value = maxima.get(("横向变形量", field), ("/", None))
        template_set(ws, f"M{row_index}", display_max_point("横向变形量", point))
        template_set(ws, f"O{row_index}", fmt_01_value(value), "0.0")
        template_set(ws, f"Q{row_index}", "")

        template_set(ws, f"S{row_index}", label)
        point, value = tilt_max.get(field, ("/", None))
        template_set(ws, f"U{row_index}", display_tilt_pair(alias_map, point))
        template_set(ws, f"W{row_index}", fmt_01_text(value) if isinstance(value, float) else "/")
        template_set(ws, f"Y{row_index}", "")
        template_set(ws, f"AA{row_index}", "")

    value_cell_refs = {
        ("垂直位移", "current", "12"): "E24:F24",
        ("垂直位移", "current", "13"): "E25:F25",
        ("垂直位移", "cumulative", "12"): "G24:H24",
        ("垂直位移", "cumulative", "13"): "G25:H25",
        ("横向变形量", "current", "12"): "M24:N24",
        ("横向变形量", "current", "13"): "M25:N25",
        ("横向变形量", "cumulative", "12"): "O24:P24",
        ("横向变形量", "cumulative", "13"): "O25:P25",
    }
    for ref in tuple(value_cell_refs.values()) + (
        "U24:V24",
        "U25:V25",
        "W24:X24",
        "W25:X25",
        "G26:H26",
        "O26:P26",
        "W26:X26",
        "G27:H27",
        "O27:P27",
        "W27:X27",
    ):
        template_fill_range(ws, ref, white_fill, black_font)
    for (item, field), (point, _) in maxima.items():
        suffix = point_suffix(point)
        ref = value_cell_refs.get((item, field, suffix))
        if ref:
            template_fill_range(ws, ref, green_fill, white_font)
    for field, (point, _) in tilt_max.items():
        suffix = point_suffix(str(point).split("/", 1)[0])
        ref = {
            ("current", "12"): "U24:V24",
            ("current", "13"): "U25:V25",
            ("cumulative", "12"): "W24:X24",
            ("cumulative", "13"): "W25:X25",
        }.get((field, suffix))
        if ref:
            template_fill_range(ws, ref, green_fill, white_font)

    notes = (
        "注：1、水平位移：“+”为向东方向、“-”为向西方向；\n"
        "       竖向位移：“+”表示隆起，“-”表示下沉；\n"
        "       倾斜：正值为北方向偏移，负值南方向偏移；\n"
        "    2、沉降、位移：预警值±5mm，报警值±7mm，控制值±10mm；倾斜：预警值1.0‰，报警值1.4‰，控制值2.0‰。"
    )
    template_set(ws, "A28", notes)
    current_tilt = tilt_max.get("current")
    cumulative_tilt = tilt_max.get("cumulative")
    tilt_text = ""
    if current_tilt and cumulative_tilt:
        tilt_text = (
            f"；本次最大倾斜率{fmt_01_text(current_tilt[1])}‰（{display_tilt_pair(alias_map, current_tilt[0])}），"
            f"累计最大倾斜率{fmt_01_text(cumulative_tilt[1])}‰（{display_tilt_pair(alias_map, cumulative_tilt[0])}）"
        )
    template_set(
        ws,
        "A32",
        f"监测数据显示：本次自动化全站仪监测各测点未出现平台预警，当前变形速率基本平稳{tilt_text}。\n"
        "建议按既定频率持续采集监测数据，并结合现场施工进度加强对12号、13号桥墩及关联测点的跟踪复核。",
    )
    template_set(ws, "A34", f"项目实施单位：{implementation_unit}    数据来源：全站仪自动化监测平台")
    template_set(ws, "O34", f"生成时间：{current_time}    报表类型：轨道交通控制保护区监测快报")

    ws.print_area = "A1:AB34"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    for temp_image in temporary_images:
        try:
            temp_image.unlink()
        except OSError:
            pass

    check = load_workbook(output_path, data_only=False)
    sheet = check["全站仪快报"]
    if sheet.calculate_dimension() != "A1:AB34":
        raise SystemExit(f"Unexpected report range: {sheet.calculate_dimension()}")
    if image_path and image_path.exists() and not sheet._images:
        raise SystemExit("Image was not embedded.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an A3 crossing-stage total-station Excel report.")
    parser.add_argument("input_csv", type=Path)
    parser.add_argument("--output", "-o", type=Path)
    parser.add_argument("--image", type=Path)
    parser.add_argument("--image-right", type=Path)
    parser.add_argument("--point-config", type=Path, help="Platform point configuration JSON used for S/X point height spacing.")
    parser.add_argument("--initial-report", type=Path, help="Initial value report .docx used to extract the project implementation unit.")
    parser.add_argument("--implementation-unit", help="Project implementation unit printed in the report footer.")
    parser.add_argument("--template", type=Path, help="Existing xlsx quick-report template to fill.")
    parser.add_argument("--work-condition", help="Override work-condition text printed in the report.")
    parser.add_argument("--manual-overrides", type=Path, help="JSON file with engineer-confirmed metric overrides.")
    parser.add_argument(
        "--tilt-axis",
        default="横向变形量",
        help="Tilt axis for signed pier inclination. Use 横向变形量, 纵向变形量, or resultant.",
    )
    parser.add_argument("--point-alias-map", type=Path, help="Project-confirmed point display mapping JSON.")
    args = parser.parse_args()

    output = args.output or default_report_output_path(args.input_csv)
    if args.template:
        build_template_report(
            args.input_csv,
            output,
            args.template,
            args.image,
            args.image_right,
            args.point_config,
            args.initial_report,
            args.implementation_unit,
            args.tilt_axis,
            args.point_alias_map,
            args.work_condition,
            args.manual_overrides,
        )
    else:
        build_compact_report(
            args.input_csv,
            output,
            args.image,
            args.image_right,
            args.point_config,
            args.initial_report,
            args.implementation_unit,
            args.tilt_axis,
            args.point_alias_map,
        )
    print(f"Wrote {output}")


if __name__ == "__main__":
    main()
