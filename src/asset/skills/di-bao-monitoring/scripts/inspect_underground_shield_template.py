#!/usr/bin/env python3
"""Inspect an underground-shield combined report workbook template.

The script does not read or transmit platform credentials. It only summarizes
workbook structure so an agent can map data sources to the engineer template
before generating a report.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_SHEETS = {"前台", "前台 (静力水准)", "Z1 全 (2)", "Z1 （静力水准)"}
KEYWORDS = [
    "工程概况",
    "施工进度",
    "各项数据分析",
    "结论与评价",
    "道床沉降",
    "结构沉降",
    "水平位移",
    "水平收敛",
    "本次最大",
    "穿越期间最大",
    "总累计最大",
]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sample_formulas(ws: Any, limit: int) -> list[dict[str, str]]:
    formulas: list[dict[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": value})
                if len(formulas) >= limit:
                    return formulas
    return formulas


def keyword_hits(ws: Any) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    patterns = {keyword: re.compile(re.escape(keyword)) for keyword in KEYWORDS}
    for row in ws.iter_rows():
        for cell in row:
            text = cell_text(cell.value)
            if not text:
                continue
            for keyword, pattern in patterns.items():
                if pattern.search(text):
                    hits.append({"keyword": keyword, "cell": cell.coordinate, "text": text[:80]})
    return hits


def worksheet_print_area(ws: Any) -> list[str]:
    value = ws.print_area
    if not value:
        return []
    if isinstance(value, str):
        return [value]
    return [str(item) for item in value]


def inspect_workbook(path: Path, formula_limit: int) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        image_count = len(getattr(ws, "_images", []) or [])
        chart_count = len(getattr(ws, "_charts", []) or [])
        sheets.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "print_area": worksheet_print_area(ws),
                "merged_cell_count": len(ws.merged_cells.ranges),
                "image_count": image_count,
                "chart_count": chart_count,
                "keyword_hits": keyword_hits(ws),
                "sample_formulas": sample_formulas(ws, formula_limit),
            }
        )
    sheet_names = {sheet["name"] for sheet in sheets}
    return {
        "workbook": str(path),
        "sheet_count": len(sheets),
        "is_underground_shield_combined_candidate": REQUIRED_SHEETS.issubset(sheet_names),
        "missing_required_sheets": sorted(REQUIRED_SHEETS - sheet_names),
        "sheets": sheets,
    }


def markdown_report(summary: dict[str, Any]) -> str:
    lines = [
        "# 地下盾构联合报表模板结构核对",
        "",
        f"- 模板：`{summary['workbook']}`",
        f"- 工作表数量：{summary['sheet_count']}",
        f"- 是否符合联合报表关键页：{'是' if summary['is_underground_shield_combined_candidate'] else '否'}",
    ]
    missing = summary["missing_required_sheets"]
    if missing:
        lines.append(f"- 缺少关键页：{', '.join(missing)}")
    lines.extend(
        [
            "",
            "## 工作表摘要",
            "",
            "| 工作表 | 显隐 | 尺寸 | 打印区域 | 合并单元格 | 图片 | 图表 | 关键词数量 | 公式样本 |",
            "|---|---|---:|---|---:|---:|---:|---:|---:|",
        ]
    )
    for sheet in summary["sheets"]:
        print_area = ", ".join(sheet["print_area"]) if sheet["print_area"] else ""
        row = dict(sheet)
        row["print_area"] = print_area
        row["hit_count"] = len(sheet["keyword_hits"])
        row["formula_count"] = len(sheet["sample_formulas"])
        lines.append(
            "| {name} | {state} | {max_row}x{max_column} | {print_area} | "
            "{merged_cell_count} | {image_count} | {chart_count} | {hit_count} | {formula_count} |".format(
                **row,
            )
        )
    lines.extend(["", "## 关键文本位置", ""])
    for sheet in summary["sheets"]:
        if not sheet["keyword_hits"]:
            continue
        lines.append(f"### {sheet['name']}")
        for hit in sheet["keyword_hits"][:40]:
            lines.append(f"- `{hit['cell']}` {hit['keyword']}：{hit['text']}")
        if len(sheet["keyword_hits"]) > 40:
            lines.append(f"- 还有 {len(sheet['keyword_hits']) - 40} 条关键词命中未展开。")
        lines.append("")
    lines.extend(["## 公式样本", ""])
    for sheet in summary["sheets"]:
        if not sheet["sample_formulas"]:
            continue
        lines.append(f"### {sheet['name']}")
        for formula in sheet["sample_formulas"]:
            lines.append(f"- `{formula['cell']}` `{formula['formula']}`")
        lines.append("")
    lines.extend(
        [
            "## 后续填报建议",
            "",
            "- 先确认沉降来源：静力水准或全站仪。",
            "- 全站仪平面变形填入 `Z1 全 (2)` 或模板对应数据页，再由 `前台` 汇总。",
            "- 静力水准沉降填入 `Z1 （静力水准)`，再由 `前台 (静力水准)` 和沉降汇总区引用。",
            "- 曲线图应继续引用数据页时间序列，不建议改成静态截图。",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("template", type=Path, help="Engineer workbook template .xlsx")
    parser.add_argument("--output", type=Path, help="Write Markdown summary to this path.")
    parser.add_argument("--json-output", type=Path, help="Optional JSON summary path.")
    parser.add_argument("--formula-limit", type=int, default=20, help="Formula samples per sheet.")
    args = parser.parse_args()

    summary = inspect_workbook(args.template, args.formula_limit)
    report = markdown_report(summary)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(report, encoding="utf-8")
    else:
        print(report, end="")
    if args.json_output:
        args.json_output.parent.mkdir(parents=True, exist_ok=True)
        args.json_output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
